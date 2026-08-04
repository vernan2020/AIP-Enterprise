from __future__ import annotations

from datetime import date
from pathlib import Path

import openpyxl
import pytest

from aip.product.configured.adapters.configured_portfolio_provider import ConfiguredPortfolioProvider
from aip.product.configured.configuration.configured_source_config import ConfiguredSourceConfig, FolderWatchSourceConfig
from aip.product.configured.readers.institutional_portfolio_master_reader import InstitutionalPortfolioMasterReader
from aip.product.demo.configuration.demo_config import DemoConfig
from aip.product.demo.configuration.environment_loader import EnvironmentLoader
from aip.tools.diagnose_configured_sources import main as diagnose_main


def test_environment_loader_defaults_configured_diagnostics_to_false(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.delenv("AIP_CONFIGURED_DIAGNOSTIC_MODE", raising=False)
    config = EnvironmentLoader().load()

    assert config.source_config["diagnostic_mode"] is False


@pytest.mark.parametrize(
    ("value", "expected"),
    [("true", True), ("1", True), ("yes", True), ("on", True), ("false", False), ("0", False), ("no", False), ("off", False), ("", False)],
)
def test_environment_loader_parses_configured_diagnostic_mode_values(monkeypatch: pytest.MonkeyPatch, value: str, expected: bool) -> None:
    monkeypatch.setenv("AIP_CONFIGURED_DIAGNOSTIC_MODE", value)
    config = EnvironmentLoader().load()

    assert config.source_config["diagnostic_mode"] is expected


def test_configured_source_config_prefers_explicit_metadata_override_over_environment(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("AIP_CONFIGURED_DIAGNOSTIC_MODE", "false")
    source_config = ConfiguredSourceConfig(diagnostic_mode=True, metadata={"diagnostic_mode": True})

    assert source_config.resolve_diagnostic_mode() is True


def test_provider_enables_diagnostics_and_omits_trace_when_disabled(tmp_path: Path) -> None:
    root = tmp_path / "institutional"
    maestro_dir = root / "Inversiones" / "2026" / "maestro" / "julio"
    vector_dir = root / "Inversiones" / "2026" / "vector" / "julio"
    maestro_dir.mkdir(parents=True)
    vector_dir.mkdir(parents=True)

    maestro_path = maestro_dir / "29-07-2026.xlsx"
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Maestro"
    worksheet.append(["ISIN", "Emisor", "Valor de Mercado", "Valor en Libros"])
    worksheet.append(["US1234567890", "Banco Central", 1000000, 980000])
    workbook.save(maestro_path)

    vector_path = vector_dir / "29-07-2026.txt"
    vector_path.write_text("BCCR  BC12M120826 12/08/2026  100.0 100.008344  2.842 0.000000 0\n", encoding="utf-8")

    config = DemoConfig(execution_mode="CONFIGURED", demo_mode_enabled=False, data_cutoff_date=date(2026, 7, 29))
    enabled_source_config = ConfiguredSourceConfig(
        folder_watch=FolderWatchSourceConfig(enabled=True, portfolio_root=str(root), vector_path=str(vector_dir)),
        diagnostic_mode=True,
    )
    disabled_source_config = ConfiguredSourceConfig(
        folder_watch=FolderWatchSourceConfig(enabled=True, portfolio_root=str(root), vector_path=str(vector_dir)),
        diagnostic_mode=False,
    )

    enabled_provider = ConfiguredPortfolioProvider(config, enabled_source_config)
    enabled_payload = enabled_provider.get_portfolio()
    disabled_provider = ConfiguredPortfolioProvider(config, disabled_source_config)
    disabled_payload = disabled_provider.get_portfolio()

    assert enabled_payload["portfolio_master"]["diagnostics"]["trace"]["records_read"] == 1
    assert enabled_payload["positions"][0]["vector_match"]["matched"] is False
    assert "record_trace" not in disabled_payload["portfolio_master"]["diagnostics"]


def test_diagnostic_cli_reports_safe_summary_and_exit_code(tmp_path: Path, monkeypatch: pytest.MonkeyPatch, capsys: pytest.CaptureFixture[str]) -> None:
    root = tmp_path / "institutional"
    maestro_dir = root / "Inversiones" / "2026" / "maestro" / "julio"
    vector_dir = root / "Inversiones" / "2026" / "vector" / "julio"
    maestro_dir.mkdir(parents=True)
    vector_dir.mkdir(parents=True)

    maestro_path = maestro_dir / "29-07-2026.xlsx"
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Maestro"
    worksheet.append(["ISIN", "Emisor", "Valor de Mercado", "Valor en Libros"])
    worksheet.append(["US1234567890", "Banco Central", 1000000, 980000])
    workbook.save(maestro_path)

    vector_path = vector_dir / "29-07-2026.txt"
    vector_path.write_text("BCCR  BC12M120826 12/08/2026  100.0 100.008344  2.842 0.000000 0\n", encoding="utf-8")

    monkeypatch.setenv("AIP_EXECUTION_MODE", "CONFIGURED")
    monkeypatch.setenv("AIP_DEMO_MODE_ENABLED", "false")
    monkeypatch.setenv("AIP_PORTFOLIO_ROOT", str(root))
    monkeypatch.setenv("AIP_VECTOR_ENABLED", "true")
    monkeypatch.setenv("AIP_VECTOR_PATH", str(vector_dir))
    monkeypatch.setenv("AIP_CONFIGURED_DIAGNOSTIC_MODE", "true")
    monkeypatch.setenv("AIP_DATA_CUTOFF_DATE", "2026-07-29")

    exit_code = diagnose_main([])
    captured = capsys.readouterr()

    assert exit_code == 0
    assert "CONFIGURED SOURCE DIAGNOSTIC" in captured.out
    assert "MASTER" in captured.out
    assert "VECTOR" in captured.out
    assert "/Inversiones" not in captured.out
    assert "US1234567890" not in captured.out


def test_diagnostic_cli_accepts_configured_mode_from_diagnostic_flag(tmp_path: Path, monkeypatch: pytest.MonkeyPatch, capsys: pytest.CaptureFixture[str]) -> None:
    root = tmp_path / "institutional"
    maestro_dir = root / "Inversiones" / "2026" / "maestro" / "julio"
    vector_dir = root / "Inversiones" / "2026" / "vector" / "julio"
    maestro_dir.mkdir(parents=True)
    vector_dir.mkdir(parents=True)

    maestro_path = maestro_dir / "29-07-2026.xlsx"
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Maestro"
    worksheet.append(["ISIN", "Emisor", "Valor de Mercado", "Valor en Libros"])
    worksheet.append(["US1234567890", "Banco Central", 1000000, 980000])
    workbook.save(maestro_path)

    vector_path = vector_dir / "29-07-2026.txt"
    vector_path.write_text("BCCR  BC12M120826 12/08/2026  100.0 100.008344  2.842 0.000000 0\n", encoding="utf-8")

    monkeypatch.delenv("AIP_EXECUTION_MODE", raising=False)
    monkeypatch.setenv("AIP_DEMO_MODE_ENABLED", "false")
    monkeypatch.setenv("AIP_PORTFOLIO_ROOT", str(root))
    monkeypatch.setenv("AIP_VECTOR_ENABLED", "true")
    monkeypatch.setenv("AIP_VECTOR_PATH", str(vector_dir))
    monkeypatch.setenv("AIP_CONFIGURED_DIAGNOSTIC_MODE", "true")
    monkeypatch.setenv("AIP_DATA_CUTOFF_DATE", "2026-07-29")

    exit_code = diagnose_main([])
    captured = capsys.readouterr()

    assert exit_code == 0
    assert "CONFIGURED SOURCE DIAGNOSTIC" in captured.out


def test_price_vector_discovery_filters_pipca_candidates_with_windows_path_and_cli(monkeypatch: pytest.MonkeyPatch, tmp_path: Path, capsys: pytest.CaptureFixture[str]) -> None:
    workspace_root = tmp_path / "Windows Root With Spaces"
    portfolio_root = workspace_root / "Inversiones" / "2026"
    vector_dir = portfolio_root / "vector" / "julio"
    maestro_dir = portfolio_root / "maestro" / "julio"
    vector_dir.mkdir(parents=True)
    maestro_dir.mkdir(parents=True)

    maestro_path = maestro_dir / "29-07-2026.xlsx"
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Maestro"
    worksheet.append(["ISIN", "Emisor", "Valor de Mercado", "Valor en Libros"])
    worksheet.append(["US1234567890", "Banco Central", 1000000, 980000])
    workbook.save(maestro_path)

    vector_path = vector_dir / "VectorPiPCA_20260729.txt"
    vector_path.write_text("BCCR  BC12M120826 12/08/2026  100.0 100.008344  2.842 0.000000 0\n", encoding="utf-8")

    monkeypatch.setenv("AIP_EXECUTION_MODE", "CONFIGURED")
    monkeypatch.setenv("AIP_DEMO_MODE_ENABLED", "false")
    monkeypatch.setenv("AIP_PORTFOLIO_ROOT", str(workspace_root))
    monkeypatch.setenv("AIP_VECTOR_ENABLED", "true")
    monkeypatch.setenv("AIP_VECTOR_PATH", str(vector_dir))
    monkeypatch.setenv("AIP_CONFIGURED_DIAGNOSTIC_MODE", "true")
    monkeypatch.setenv("AIP_DATA_CUTOFF_DATE", "2026-07-29")

    config = EnvironmentLoader().load()
    provider = ConfiguredPortfolioProvider(config, ConfiguredSourceConfig.from_safe_dict(config.source_config) if hasattr(ConfiguredSourceConfig, "from_safe_dict") else None)
    payload = provider.get_portfolio()

    assert payload["price_vector"]["file_name"] == "VectorPiPCA_20260729.txt"
    assert payload["price_vector"]["valuation_date"] == "2026-07-29"
    assert payload["price_vector"]["status"] == "HEALTHY"
    assert payload["price_vector"]["diagnostics"]["pipca_candidate_count"] == 1
    assert payload["price_vector"]["diagnostics"]["exact_date_match_count"] == 1
    assert payload["price_vector"]["diagnostics"]["candidate_count"] == 1

    exit_code = diagnose_main([])
    captured = capsys.readouterr()

    assert exit_code == 0
    assert "VectorPiPCA_20260729.txt" in captured.out
    assert "Status: HEALTHY" in captured.out


def test_master_reader_limits_detailed_trace_output(tmp_path: Path) -> None:
    master_path = tmp_path / "master.xlsx"
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Maestro"
    worksheet.append(["ISIN", "Emisor", "Valor de Mercado", "Valor en Libros"])
    for index in range(25):
        worksheet.append([f"ISIN{index}", f"Issuer {index}", 1000 + index, 900 + index])
    workbook.save(master_path)

    reader = InstitutionalPortfolioMasterReader()
    result = reader.read(master_path, diagnostic_mode=True)

    assert len(result.diagnostics["trace"]["record_trace"]) == 20
    assert result.diagnostics["trace"]["record_trace"][0]["row"] >= 2
    assert result.diagnostics["trace"]["record_trace"][-1]["row"] >= 22
