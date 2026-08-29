from __future__ import annotations

import os
import subprocess
import sys
from datetime import date
from decimal import Decimal
from pathlib import Path

import openpyxl

from aip.product.configured.adapters.configured_portfolio_provider import (
    ConfiguredPortfolioProvider,
)
from aip.product.configured.configuration.configured_source_config import (
    ConfiguredSourceConfig,
    FolderWatchSourceConfig,
    VectorSourceConfig,
)
from aip.product.configured.readers.pipca_vector_reader import InstitutionalPiPCAVectorReader
from aip.product.configured.services.institutional_matching_service import (
    InstitutionalPortfolioMatchingService,
)
from aip.product.demo.configuration.demo_config import DemoConfig
from aip.tools.reconcile_portfolio_valuation import _build_reconciliation_rows


def _build_configured_provider(
    tmp_path: Path, *, row_values: list[list[object]]
) -> ConfiguredPortfolioProvider:
    master_root = tmp_path / "institutional" / "Inversiones" / "2026"
    master_dir = master_root / "maestro"
    master_dir.mkdir(parents=True, exist_ok=True)
    workbook_path = master_dir / "29-07-2026.xlsx"
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append(
        [
            "issuer",
            "product_code",
            "series",
            "maturity_date",
            "market_value",
            "book_value",
            "isin",
            "currency",
            "valor mercado colonizado",
        ]
    )
    for row in row_values:
        worksheet.append(row)
    workbook.save(workbook_path)

    config = DemoConfig(
        execution_mode="CONFIGURED", demo_mode_enabled=False, data_cutoff_date=date(2026, 7, 29)
    )
    source_config = ConfiguredSourceConfig(
        folder_watch=FolderWatchSourceConfig(
            enabled=True, portfolio_root=str(tmp_path / "institutional")
        ),
        vector=VectorSourceConfig(enabled=True, path=str(tmp_path / "vector")),
        metadata={
            "allow_prior_source_date": False,
            "data_cutoff_date": date(2026, 7, 29).isoformat(),
            "diagnostic_mode": "false",
        },
    )
    return ConfiguredPortfolioProvider(config, source_config)


def _build_pipca_line(
    series: str, *, issuer: str = "BCCR", mnemonic: str = "BEM", maturity: str = "24/03/2027"
) -> str:
    return f"{issuer}{mnemonic:<6}{series}{maturity}".ljust(120)


def test_reconciliation_summary_classifies_positions_and_reports_collisions() -> None:
    service = InstitutionalPortfolioMatchingService()
    master_positions = [
        {
            "isin": "",
            "series": "S240327",
            "maturity_date": date(2027, 3, 24),
            "issuer": "Banco Central",
            "product_code": "tpras",
            "market_value": 1000.0,
        },
        {
            "isin": "",
            "series": "B180429",
            "maturity_date": date(2029, 4, 18),
            "issuer": "Banco Central",
            "product_code": "tptba",
            "market_value": 2000.0,
        },
        {
            "isin": "",
            "series": "X999999",
            "maturity_date": date(2028, 1, 1),
            "issuer": "Tesoreria",
            "product_code": "tpran",
            "market_value": 3000.0,
        },
        {
            "isin": "",
            "series": "FUND01",
            "maturity_date": None,
            "issuer": "Banco Central",
            "product_code": "fund",
            "market_value": 4000.0,
        },
    ]
    vector_records = [
        {
            "issuer": "G",
            "instrument_type_or_mnemonic": "tpras",
            "series_or_security_code": "S240327",
            "maturity_date_if_present": date(2027, 3, 24),
            "market_price": Decimal("99.5"),
            "market_yield": Decimal("6.0"),
            "source_index": 0,
        },
        {
            "issuer": "G",
            "instrument_type_or_mnemonic": "tptba",
            "series_or_security_code": "B180429",
            "maturity_date_if_present": date(2029, 4, 18),
            "market_price": Decimal("100.0"),
            "market_yield": Decimal("5.5"),
            "source_index": 1,
        },
        {
            "issuer": "G",
            "instrument_type_or_mnemonic": "tpras",
            "series_or_security_code": "S240327",
            "maturity_date_if_present": date(2027, 3, 24),
            "market_price": Decimal("99.5"),
            "market_yield": Decimal("6.0"),
            "source_index": 2,
        },
    ]

    enriched_positions, summary = service.enrich_positions(master_positions, vector_records)
    reconciliation = summary["reconciliation"]

    assert reconciliation["total_master_positions"] == 4
    assert reconciliation["eligible_fixed_income_positions"] == 3
    assert reconciliation["matched_positions"] == 2
    assert reconciliation["expected_unmatched_positions"] == 1
    assert reconciliation["review_required_positions"] == 1
    assert reconciliation["raw_match_percentage"] == 50.0
    assert reconciliation["eligible_match_percentage"] == 66.7
    assert reconciliation["vector_key_collisions"] == 1
    assert reconciliation["collisions_affecting_portfolio"] == 1
    assert reconciliation["unmatched_reason_groups"]["series absent from vector"] == 1
    assert reconciliation["position_reconciliation"][0]["classification"] == "MATCHED"
    assert (
        reconciliation["position_reconciliation"][3]["reason_no_match"]
        == "no maturity / perpetual / fund"
    )
    assert reconciliation["known_government_positions"]["S240327"]["classification"] == "MATCHED"


def test_exact_series_plus_maturity_match_uses_date_objects() -> None:
    service = InstitutionalPortfolioMatchingService()
    master_positions = [
        {
            "isin": "",
            "series": "S240327",
            "maturity_date": date(2027, 3, 24),
            "issuer": "Banco Central",
            "product_code": "BEM",
        }
    ]
    vector_records = [
        {
            "issuer": "Banco Central",
            "instrument_type_or_mnemonic": "BEM",
            "series_or_security_code": "S240327",
            "maturity_date_if_present": date(2027, 3, 24),
            "source_index": 0,
        }
    ]

    enriched_positions, summary = service.enrich_positions(master_positions, vector_records)

    assert summary["series_maturity_matches"] == 1
    assert enriched_positions[0]["vector_match"]["matched"] is True
    assert enriched_positions[0]["vector_match"]["match_method"] == "SERIES_MATURITY"
    assert (
        enriched_positions[0]["matching_diagnostics"]["matching_keys"]["series_maturity"]
        == "s240327|2027-03-24"
    )


def test_leading_zero_series_is_preserved_by_positional_parser() -> None:
    reader = InstitutionalPiPCAVectorReader()
    record = reader._parse_line(
        _build_pipca_line("001234", maturity="24/03/2027"),
        source_cutoff=date(2026, 7, 29),
        source_line=1,
    )

    assert record is not None
    assert record.series_or_security_code == "001234"
    assert record.normalized_series_key == "001234"


def test_adjacent_fixed_width_fields_are_parsed_without_losing_series() -> None:
    reader = InstitutionalPiPCAVectorReader()
    record = reader._parse_line(
        _build_pipca_line("S240327", maturity="24/03/2027"),
        source_cutoff=date(2026, 7, 29),
        source_line=1,
    )

    assert record is not None
    assert record.issuer == "BCCR"
    assert record.instrument_type_or_mnemonic == "BEM"
    assert record.series_or_security_code == "S240327"
    assert record.normalized_issuer_key == "bccr"
    assert record.normalized_series_key == "s240327"


def test_costa_rican_government_security_series_is_preserved() -> None:
    reader = InstitutionalPiPCAVectorReader()
    record = reader._parse_line(
        _build_pipca_line("CRS240129", maturity="24/01/2029"),
        source_cutoff=date(2026, 7, 29),
        source_line=1,
    )

    assert record is not None
    assert record.series_or_security_code == "CRS240129"
    assert record.normalized_series_key == "crs240129"
    assert record.maturity_date_if_present == date(2029, 1, 24)


def test_production_pipca_combined_product_series_layouts_are_split_correctly() -> None:
    reader = InstitutionalPiPCAVectorReader()
    samples = [
        ("tptba", "B180429", "18/04/2029", "tptbaB180429"),
        ("tpras", "S240327", "24/03/2027", "tprasS240327"),
        ("tpras", "CRS240129", "24/01/2029", "tprasCRS240129"),
    ]

    for product_code, series_code, maturity_text, raw_field in samples:
        line = f"BCCR{product_code}{series_code}{maturity_text}".ljust(120)
        record = reader._parse_line(line, source_cutoff=date(2026, 7, 29), source_line=1)

        print(f"raw: {raw_field}")
        print(f"product: {record.instrument_type_or_mnemonic if record is not None else None}")
        print(f"series: {record.series_or_security_code if record is not None else None}")
        print(f"maturity: {maturity_text}")
        print(f"issuer: {record.issuer if record is not None else None}")

        assert record is not None
        assert record.issuer == "BCCR"
        assert record.instrument_type_or_mnemonic == product_code
        assert record.series_or_security_code == series_code
        if maturity_text == "18/04/2029":
            assert record.maturity_date_if_present == date(2029, 4, 18)
        elif maturity_text == "24/03/2027":
            assert record.maturity_date_if_present == date(2027, 3, 24)
        else:
            assert record.maturity_date_if_present == date(2029, 1, 24)


def test_public_reader_accepts_production_fixed_width_lines_and_matches_by_series_maturity(
    tmp_path,
) -> None:
    path = tmp_path / "VectorPiPCA_20260729.txt"
    path.write_text(
        "G   tptbaB180429   18/04/2029\n"
        "G   tprasS240327   24/03/2027\n"
        "G   tprasCRS240129 24/01/2029\n",
        encoding="utf-8",
    )

    result = InstitutionalPiPCAVectorReader().read(
        path, source_cutoff=date(2026, 7, 29), diagnostic_mode=True
    )

    assert result.accepted_count == 3
    assert result.rejected_count == 0
    assert [record.issuer for record in result.records] == ["G", "G", "G"]
    assert [record.instrument_type_or_mnemonic for record in result.records] == [
        "tptba",
        "tpras",
        "tpras",
    ]
    assert [record.series_or_security_code for record in result.records] == [
        "B180429",
        "S240327",
        "CRS240129",
    ]
    assert [record.maturity_date_if_present for record in result.records] == [
        date(2029, 4, 18),
        date(2027, 3, 24),
        date(2029, 1, 24),
    ]

    service = InstitutionalPortfolioMatchingService()
    records_for_matching = [
        {
            "issuer": record.issuer,
            "instrument_type_or_mnemonic": record.instrument_type_or_mnemonic,
            "series_or_security_code": record.series_or_security_code,
            "maturity_date_if_present": record.maturity_date_if_present,
            "source_line": record.source_line,
        }
        for record in result.records
    ]
    enriched_positions, summary = service.enrich_positions(
        [
            {
                "isin": "",
                "series": "B180429",
                "maturity_date": date(2029, 4, 18),
                "issuer": "Banco Central",
                "product_code": "tptba",
            }
        ],
        records_for_matching,
    )

    assert summary["series_maturity_matches"] == 1
    assert enriched_positions[0]["vector_match"]["matched"] is True
    assert enriched_positions[0]["vector_match"]["match_method"] == "SERIES_MATURITY"


def test_provider_pipeline_matches_real_pipca_line(tmp_path: Path) -> None:
    root = tmp_path / "institutional" / "Inversiones" / "2026"
    (root / "maestro").mkdir(parents=True)
    (root / "vector").mkdir(parents=True)

    workbook_path = root / "maestro" / "29-07-2026.xlsx"
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append(
        ["issuer", "product_code", "series", "maturity_date", "market_value", "book_value", "isin"]
    )
    worksheet.append(["Banco Central", "tptba", "B180429", "2029-04-18", 1000.0, 1000.0, ""])
    workbook.save(workbook_path)

    vector_path = root / "vector" / "VectorPiPCA_20260729.txt"
    vector_path.write_text("G   tptbaB180429   18/04/2029\n", encoding="utf-8")

    config = DemoConfig(
        execution_mode="CONFIGURED", demo_mode_enabled=False, data_cutoff_date=date(2026, 7, 29)
    )
    source_config = ConfiguredSourceConfig(
        folder_watch=FolderWatchSourceConfig(
            enabled=True, portfolio_root=str(tmp_path / "institutional")
        ),
        metadata={
            "allow_prior_source_date": False,
            "data_cutoff_date": date(2026, 7, 29).isoformat(),
            "diagnostic_mode": "true",
        },
    )
    provider = ConfiguredPortfolioProvider(config, source_config)

    payload = provider.get_portfolio()
    positions = payload["positions"]

    assert payload["price_vector"]["status"] == "HEALTHY"
    assert positions[0]["series"] == "B180429"
    assert positions[0]["product_code"] == "tptba"
    assert positions[0]["maturity_date"] == date(2029, 4, 18)
    assert positions[0]["vector_match"]["matched"] is True
    assert positions[0]["vector_match"]["match_method"] == "SERIES_MATURITY"


def test_inspect_pipca_vector_cli_search_reports_accepted_production_lines(
    tmp_path, capsys
) -> None:
    path = tmp_path / "VectorPiPCA_20260729.txt"
    path.write_text(
        "G   tptbaB180429   18/04/2029\n"
        "G   tprasS240327   24/03/2027\n"
        "G   tprasCRS240129 24/01/2029\n",
        encoding="utf-8",
    )

    import sys

    from aip.tools.inspect_pipca_vector import main

    sys.argv = ["inspect_pipca_vector", str(path), "--search", "B180429"]
    main()

    captured = capsys.readouterr()
    assert "status=accepted" in captured.out
    assert "raw_identifier" in captured.out


def test_inspect_pipca_vector_cli_search_reports_rejected_production_lines(
    tmp_path, capsys
) -> None:
    path = tmp_path / "VectorPiPCA_20260729.txt"
    path.write_text(
        "G   tptbaB180429\n" "G   tprasS240327   24/03/2027\n" "G   tprasCRS240129 24/01/2029\n",
        encoding="utf-8",
    )

    import sys

    from aip.tools.inspect_pipca_vector import main

    sys.argv = ["inspect_pipca_vector", str(path), "--search", "B180429"]
    main()

    captured = capsys.readouterr()
    assert "status=rejected" in captured.out
    assert "reason=missing maturity" in captured.out
    assert "branch=maturity-scan" in captured.out
    assert "parsed_series=B180429" in captured.out


def test_inspect_pipca_vector_cli_search_reports_no_matching_raw_lines(tmp_path, capsys) -> None:
    path = tmp_path / "VectorPiPCA_20260729.txt"
    path.write_text(
        "G   tptbaB180429   18/04/2029\n",
        encoding="utf-8",
    )

    import sys

    from aip.tools.inspect_pipca_vector import main

    sys.argv = ["inspect_pipca_vector", str(path), "--search", "NOT_PRESENT"]
    main()

    captured = capsys.readouterr()
    assert "No raw lines contained token 'NOT_PRESENT'" in captured.out


def test_zero_isin_vector_matches_through_secondary_keys() -> None:
    service = InstitutionalPortfolioMatchingService()
    master_positions = [
        {
            "isin": "",
            "series": "B180429",
            "maturity_date": date(2029, 4, 18),
            "issuer": "Banco Central",
            "product_code": "BEM",
        }
    ]
    vector_records = [
        {
            "issuer": "Banco Central",
            "instrument_type_or_mnemonic": "BEM",
            "series_or_security_code": "B180429",
            "maturity_date_if_present": date(2029, 4, 18),
            "isin_if_present": "",
            "source_index": 0,
        }
    ]

    enriched_positions, summary = service.enrich_positions(master_positions, vector_records)

    assert summary["exact_isin_matches"] == 0
    assert summary["series_maturity_matches"] == 1
    assert enriched_positions[0]["vector_match"]["matched"] is True
    assert enriched_positions[0]["vector_match"]["match_method"] == "SERIES_MATURITY"


def test_ambiguous_secondary_matches_are_not_silently_selected() -> None:
    service = InstitutionalPortfolioMatchingService()
    master_positions = [
        {
            "isin": "",
            "series": "S240327",
            "maturity_date": date(2027, 3, 24),
            "issuer": "Banco Central",
            "product_code": "BEM",
        }
    ]
    vector_records = [
        {
            "issuer": "Banco Central",
            "instrument_type_or_mnemonic": "BEM",
            "series_or_security_code": "S240327",
            "maturity_date_if_present": date(2027, 3, 24),
            "source_index": 0,
        },
        {
            "issuer": "Banco Central",
            "instrument_type_or_mnemonic": "BEM",
            "series_or_security_code": "S240327",
            "maturity_date_if_present": date(2027, 3, 24),
            "source_index": 1,
        },
    ]

    enriched_positions, summary = service.enrich_positions(master_positions, vector_records)

    assert summary["ambiguous_matches"] == 1
    assert enriched_positions[0]["vector_match"]["matched"] is False
    assert enriched_positions[0]["vector_match"]["ambiguous"] is True
    assert enriched_positions[0]["vector_match"]["match_method"] == "NO_VECTOR_MATCH"


def test_configured_provider_uses_colonized_market_value_for_crc_positions(tmp_path: Path) -> None:
    provider = _build_configured_provider(
        tmp_path,
        row_values=[
            ["Banco Central", "tptba", "S240327", "2027-03-24", 1500.0, 1400.0, "", "CRC", 1500.0]
        ],
    )

    payload = provider.get_portfolio()
    position = payload["positions"][0]

    assert position["market_value_crc"] == 1500.0
    assert position["market_value_local"] == 1500.0
    assert payload["market_value"] == 1500.0


def test_configured_provider_uses_colonized_market_value_for_usd_positions(tmp_path: Path) -> None:
    provider = _build_configured_provider(
        tmp_path,
        row_values=[
            ["Banco Central", "tptba", "B180429", "2029-04-18", 2500.0, 2400.0, "", "USD", 180000.0]
        ],
    )

    payload = provider.get_portfolio()
    position = payload["positions"][0]

    assert position["currency"] == "USD"
    assert position["market_value_local"] == 2500.0
    assert position["market_value_crc"] == 180000.0
    assert payload["market_value"] == 180000.0


def test_configured_provider_falls_back_when_colonized_value_is_missing(tmp_path: Path) -> None:
    provider = _build_configured_provider(
        tmp_path,
        row_values=[
            ["Banco Central", "tptba", "CRS240129", "2029-01-24", 750.0, 700.0, "", "CRC", None]
        ],
    )

    payload = provider.get_portfolio()
    position = payload["positions"][0]

    assert position["market_value_local"] == 750.0
    assert position["market_value_crc"] == 750.0
    assert payload["market_value"] == 750.0


def test_configured_provider_preserves_match_status_for_reconciliation(tmp_path: Path) -> None:
    master_root = tmp_path / "institutional" / "Inversiones" / "2026"
    vector_root = tmp_path / "vector"
    vector_root.mkdir(parents=True, exist_ok=True)
    (master_root / "maestro").mkdir(parents=True, exist_ok=True)
    workbook_path = master_root / "maestro" / "29-07-2026.xlsx"
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append(
        [
            "issuer",
            "product_code",
            "series",
            "maturity_date",
            "market_value",
            "book_value",
            "isin",
            "currency",
        ]
    )
    worksheet.append(["Banco Central", "tptba", "B180429", "2029-04-18", 1000.0, 1000.0, "", "USD"])
    workbook.save(workbook_path)

    vector_path = vector_root / "VectorPiPCA_20260729.txt"
    vector_path.write_text("G   tptbaB180429   18/04/2029\n", encoding="utf-8")

    config = DemoConfig(
        execution_mode="CONFIGURED", demo_mode_enabled=False, data_cutoff_date=date(2026, 7, 29)
    )
    source_config = ConfiguredSourceConfig(
        folder_watch=FolderWatchSourceConfig(
            enabled=True, portfolio_root=str(tmp_path / "institutional")
        ),
        vector=VectorSourceConfig(enabled=True, path=str(vector_root)),
        metadata={
            "allow_prior_source_date": False,
            "data_cutoff_date": date(2026, 7, 29).isoformat(),
            "diagnostic_mode": "false",
        },
    )
    provider = ConfiguredPortfolioProvider(config, source_config)

    payload = provider.get_portfolio()
    row = _build_reconciliation_rows(payload["positions"])[0]

    assert payload["positions"][0]["match_status"] == "MATCHED"
    assert row["matched_status"] == "MATCHED"


def test_diagnose_configured_sources_cli_matches_production_pipca_line(tmp_path: Path) -> None:
    repo_root = Path(__file__).resolve().parents[4]
    portfolio_root = tmp_path / "portfolio"
    investment_root = portfolio_root / "Inversiones"
    master_dir = investment_root / "2026" / "maestro" / "julio"
    master_dir.mkdir(parents=True)

    workbook_path = master_dir / "29-07-2026.xlsx"
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append(
        ["issuer", "product_code", "series", "maturity_date", "market_value", "book_value", "isin"]
    )
    worksheet.append(["Banco Central", "tptba", "B180429", "2029-04-18", 1000.0, 1000.0, ""])
    workbook.save(workbook_path)

    vector_dir = tmp_path / "vector"
    vector_dir.mkdir(parents=True)
    vector_path = vector_dir / "VectorPiPCA_20260729.txt"
    vector_path.write_text("G   tptbaB180429   18/04/2029\n", encoding="utf-8")

    env = os.environ.copy()
    env.update(
        {
            "AIP_EXECUTION_MODE": "CONFIGURED",
            "AIP_CONFIGURED_DIAGNOSTIC_MODE": "true",
            "AIP_FOLDERWATCH_ENABLED": "true",
            "AIP_VECTOR_ENABLED": "true",
            "AIP_PORTFOLIO_ROOT": str(portfolio_root),
            "AIP_VECTOR_PATH": str(vector_dir),
            "AIP_DATA_CUTOFF_DATE": "2026-07-29",
        }
    )
    env["PYTHONPATH"] = str(repo_root / "src") + os.pathsep + env.get("PYTHONPATH", "")

    completed = subprocess.run(
        [sys.executable, "-m", "aip.tools.diagnose_configured_sources"],
        cwd=repo_root,
        env=env,
        capture_output=True,
        text=True,
        check=False,
    )

    assert completed.returncode == 0, completed.stdout + completed.stderr
    assert "Matched by series/maturity: 1" in completed.stdout
    assert "vector_records_available_for_matching: 1" in completed.stdout
    assert "vector_key_sample" in completed.stdout
    assert "master_key_sample" in completed.stdout
