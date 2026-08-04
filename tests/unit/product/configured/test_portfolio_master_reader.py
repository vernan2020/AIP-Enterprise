from __future__ import annotations

from datetime import date
from pathlib import Path

import openpyxl
import pytest
import xlwt

from aip.product.configured.adapters.configured_portfolio_provider import ConfiguredPortfolioProvider
from aip.product.configured.configuration.configured_source_config import ConfiguredSourceConfig, FolderWatchSourceConfig
from aip.product.configured.readers.portfolio_master_reader import PortfolioMasterReader
from aip.product.demo.configuration.demo_config import DemoConfig


def _write_xlsx(path: Path, *, sheet_name: str = "Maestro") -> None:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(["Resumen institucional", ""])
    worksheet.append(["", ""])
    worksheet.append(["ISIN", "Emisor", "Instrumento", "Moneda", "Nominal", "Valor de Mercado", "Valor en Libros", "Fecha de Vencimiento"])
    worksheet.append(["US1234567890", "Banco Central", "Bono", "USD", "1000000", "1005000", "995000", date(2028, 1, 15)])
    workbook.save(path)


def _write_xls(path: Path) -> None:
    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet("Maestro")
    sheet.write(0, 0, "Resumen institucional")
    sheet.write(2, 0, "ISIN")
    sheet.write(2, 1, "Emisor")
    sheet.write(2, 2, "Instrumento")
    sheet.write(2, 3, "Moneda")
    sheet.write(2, 4, "Nominal")
    sheet.write(2, 5, "Valor de Mercado")
    sheet.write(2, 6, "Valor en Libros")
    sheet.write(2, 7, "Fecha de Vencimiento")
    sheet.write(3, 0, "US1234567890")
    sheet.write(3, 1, "Banco Central")
    sheet.write(3, 2, "Bono")
    sheet.write(3, 3, "USD")
    sheet.write(3, 4, 1000000)
    sheet.write(3, 5, 1005000)
    sheet.write(3, 6, 995000)
    sheet.write(3, 7, date(2028, 1, 15))
    workbook.save(str(path))


def test_portfolio_master_reader_reads_xlsx_with_header_detection_and_normalization(tmp_path: Path) -> None:
    path = tmp_path / "maestro.xlsx"
    _write_xlsx(path)

    result = PortfolioMasterReader().read(path, valuation_date_override=date(2026, 7, 29))

    assert result.source_status == "HEALTHY"
    assert result.sheet_selected == "Maestro"
    assert result.valuation_date == date(2026, 7, 29)
    assert result.rejected_row_count == 0
    assert result.normalized_positions[0]["isin"] == "US1234567890"
    assert result.normalized_positions[0]["market_value"] == 1005000
    assert result.normalized_positions[0]["book_value"] == 995000


def test_portfolio_master_reader_reads_xls_and_marks_duplicates_and_malformed_rows(tmp_path: Path) -> None:
    path = tmp_path / "maestro.xls"
    _write_xls(path)

    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet("Maestro")
    sheet.write(0, 0, "Resumen")
    sheet.write(2, 0, "ISIN")
    sheet.write(2, 1, "Emisor")
    sheet.write(2, 2, "Instrumento")
    sheet.write(2, 3, "Moneda")
    sheet.write(2, 4, "Nominal")
    sheet.write(2, 5, "Valor de Mercado")
    sheet.write(2, 6, "Valor en Libros")
    sheet.write(3, 0, "US1234567890")
    sheet.write(3, 1, "Banco Central")
    sheet.write(3, 2, "Bono")
    sheet.write(3, 3, "USD")
    sheet.write(3, 4, 1000000)
    sheet.write(3, 5, 1005000)
    sheet.write(3, 6, 995000)
    sheet.write(4, 0, "US1234567890")
    sheet.write(4, 1, "Banco Central")
    sheet.write(4, 2, "Bono")
    sheet.write(4, 3, "USD")
    sheet.write(4, 4, 1000000)
    sheet.write(4, 5, 1005000)
    sheet.write(4, 6, 995000)
    sheet.write(5, 0, "")
    sheet.write(5, 1, "Incomplete")
    sheet.write(5, 2, "")
    sheet.write(5, 3, "")
    workbook.save(str(path))

    result = PortfolioMasterReader().read(path, valuation_date_override=date(2026, 7, 29))

    assert result.source_status == "DEGRADED"
    assert result.rejected_row_count >= 1
    assert result.warnings
    assert len(result.normalized_positions) == 1


def test_configured_provider_returns_real_positions_without_demo_fallback(tmp_path: Path) -> None:
    root = tmp_path / "institutional"
    path = root / "Inversiones" / "2026" / "maestro" / "julio" / "29-07-2026.xlsx"
    path.parent.mkdir(parents=True)
    _write_xlsx(path)

    config = DemoConfig(execution_mode="CONFIGURED", demo_mode_enabled=False, data_cutoff_date=date(2026, 7, 29))
    source_config = ConfiguredSourceConfig(folder_watch=FolderWatchSourceConfig(enabled=True, portfolio_root=str(root)))
    provider = ConfiguredPortfolioProvider(config, source_config)

    payload = provider.get_portfolio()

    assert payload["positions"]
    assert payload["market_value"] > 0
    assert payload["book_value"] > 0
    assert payload["portfolio_master"]["file_name"] == "29-07-2026.xlsx"
    assert "Acme Bank" not in str(payload)
    assert "Blue Ridge" not in str(payload)
