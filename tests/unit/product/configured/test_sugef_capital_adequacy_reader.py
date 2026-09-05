from __future__ import annotations

import io
from datetime import date
from decimal import Decimal

from openpyxl import Workbook

from aip.product.configured.configuration.configured_source_config import (
    SUGEFFinancialSourceConfig,
)
from aip.product.configured.readers.sugef_capital_adequacy_reader import (
    SUGEFCapitalAdequacyReader,
)
from aip.product.configured.readers.sugef_public_api_client import (
    SUGEFPublicApiResponse,
)


class _EntityApi:
    def list_entities(self) -> SUGEFPublicApiResponse:
        rows = (
            {
                "codigoEntidad": "3004045138",
                "nombreEntidad": "COOPEALIANZA R.L.",
                "aliasEntidad": "COOPEALIANZA",
                "aliasPublicacionEntidad": "COOPEALIANZA R.L.",
                "descripcionSector": "Cooperativas",
            },
        )
        return SUGEFPublicApiResponse(
            operation="/Catalogo/MAPI/ListarEntidades",
            endpoint="https://sugef.example/entities",
            method="GET",
            body={"listaEntidades": list(rows)},
            rows=rows,
        )


def _workbook_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Suficiencia"
    sheet.append(("Entidad", "Suficiencia Patrimonial (%)"))
    sheet.append(("COOPEALIANZA R.L.", 20.25))
    payload = io.BytesIO()
    workbook.save(payload)
    return payload.getvalue()


def _multilevel_workbook_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Suficiencia Patrimonial"
    sheet.append(("Reporte del Indicador de Suficiencia Patrimonial", None))
    sheet.append(("Sector / Entidad supervisada", "Indicador"))
    sheet.append((None, "Suficiencia Patrimonial (%)"))
    sheet.append(("COOPEALIANZA R.L. - Cooperativa de Ahorro y Crédito", 20.25))
    payload = io.BytesIO()
    workbook.save(payload)
    return payload.getvalue()


def _historical_ispe_workbook_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Histórico"
    sheet.append(("Entidad", "Fecha de Corte", "ISPE"))
    sheet.append(("COOPEALIANZA R.L.", date(2026, 3, 31), 19.75))
    sheet.append(("COOPEALIANZA R.L.", date(2026, 6, 30), 20.25))
    payload = io.BytesIO()
    workbook.save(payload)
    return payload.getvalue()


def _inferred_flat_workbook_bytes(*, ambiguous: bool = False) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Histórico"
    if ambiguous:
        sheet.append(("codigo", "corte_reportado", "valor_a", "valor_b"))
        sheet.append(("3004045138", date(2026, 6, 30), 20.25, 14.75))
    else:
        sheet.append(("codigo", "corte_reportado", "valor"))
        sheet.append(("3004045138", date(2026, 6, 30), 20.25))
    payload = io.BytesIO()
    workbook.save(payload)
    return payload.getvalue()


def _page() -> bytes:
    return b"""
    <html><body>
      <a href="/reportes/sp/Suficiencia%20Patrimonial%20(Junio%202026).xlsx">Junio 2026</a>
    </body></html>
    """


def test_reader_uses_latest_quarterly_cutoff_and_carries_it_to_analysis_date() -> None:
    workbook = _workbook_bytes()
    page = b"""
    <html><body>
      <a href="/reportes/sp/Suficiencia%20Patrimonial%20(Marzo%202026).xlsx">Marzo 2026</a>
      <a href="/reportes/sp/Suficiencia%20Patrimonial%20(Junio%202026).xlsx">Junio 2026</a>
    </body></html>
    """

    def fetch(url: str) -> bytes:
        if url == SUGEFCapitalAdequacyReader.SOURCE_PAGE:
            return page
        return workbook

    reader = SUGEFCapitalAdequacyReader(
        SUGEFFinancialSourceConfig(),
        api_client=_EntityApi(),  # type: ignore[arg-type]
        fetch_bytes=fetch,
    )

    result = reader.read(date(2026, 7, 31))

    assert result.source_cutoff == date(2026, 6, 30)
    assert len(result.lines) == 1
    line = result.lines[0]
    assert line.statement_date == date(2026, 7, 31)
    assert line.account_code == "SUGEF:CAPITAL_ADEQUACY"
    assert line.amount == Decimal("0.2025")
    assert line.trace is not None
    assert "30/06/2026" in line.trace.file_path
    assert any("último corte trimestral oficial aplicable" in item for item in result.diagnostics)


def test_reader_supports_multilevel_header_and_unique_extended_entity_name() -> None:
    workbook = _multilevel_workbook_bytes()

    def fetch(url: str) -> bytes:
        if url == SUGEFCapitalAdequacyReader.SOURCE_PAGE:
            return _page()
        return workbook

    reader = SUGEFCapitalAdequacyReader(
        SUGEFFinancialSourceConfig(),
        api_client=_EntityApi(),  # type: ignore[arg-type]
        fetch_bytes=fetch,
    )

    result = reader.read(date(2026, 7, 31))

    assert result.source_cutoff == date(2026, 6, 30)
    assert len(result.lines) == 1
    assert result.lines[0].entity.entity_id == "3004045138"
    assert result.lines[0].amount == Decimal("0.2025")
    assert any("1 entidades cargadas" in item for item in result.diagnostics)


def test_reader_supports_official_ispe_header_and_filters_historical_cutoff() -> None:
    workbook = _historical_ispe_workbook_bytes()

    def fetch(url: str) -> bytes:
        if url == SUGEFCapitalAdequacyReader.SOURCE_PAGE:
            return _page()
        return workbook

    reader = SUGEFCapitalAdequacyReader(
        SUGEFFinancialSourceConfig(),
        api_client=_EntityApi(),  # type: ignore[arg-type]
        fetch_bytes=fetch,
    )

    result = reader.read(date(2026, 7, 31))

    assert len(result.lines) == 1
    line = result.lines[0]
    assert line.amount == Decimal("0.2025")
    assert line.trace is not None
    assert line.trace.sheet_name == "Histórico"
    assert line.trace.row_number == 3
    assert any("encabezados explícitos/ISPE" in item for item in result.diagnostics)


def test_reader_can_infer_unambiguous_flat_historical_layout() -> None:
    workbook = _inferred_flat_workbook_bytes()

    def fetch(url: str) -> bytes:
        if url == SUGEFCapitalAdequacyReader.SOURCE_PAGE:
            return _page()
        return workbook

    reader = SUGEFCapitalAdequacyReader(
        SUGEFFinancialSourceConfig(),
        api_client=_EntityApi(),  # type: ignore[arg-type]
        fetch_bytes=fetch,
    )

    result = reader.read(date(2026, 7, 31))

    assert len(result.lines) == 1
    assert result.lines[0].amount == Decimal("0.2025")
    assert any("inferencia estructural controlada" in item for item in result.diagnostics)


def test_reader_refuses_ambiguous_inferred_numeric_columns() -> None:
    workbook = _inferred_flat_workbook_bytes(ambiguous=True)

    def fetch(url: str) -> bytes:
        if url == SUGEFCapitalAdequacyReader.SOURCE_PAGE:
            return _page()
        return workbook

    reader = SUGEFCapitalAdequacyReader(
        SUGEFFinancialSourceConfig(),
        api_client=_EntityApi(),  # type: ignore[arg-type]
        fetch_bytes=fetch,
    )

    result = reader.read(date(2026, 7, 31))

    assert result.lines == ()
    assert any("no se identificaron observaciones inequívocas" in item for item in result.diagnostics)


def test_reader_does_not_use_future_quarter() -> None:
    workbook = _workbook_bytes()

    def fetch(url: str) -> bytes:
        if url == SUGEFCapitalAdequacyReader.SOURCE_PAGE:
            return _page()
        return workbook

    reader = SUGEFCapitalAdequacyReader(
        SUGEFFinancialSourceConfig(),
        api_client=_EntityApi(),  # type: ignore[arg-type]
        fetch_bytes=fetch,
    )

    result = reader.read(date(2026, 5, 31))

    assert result.lines == ()
    assert result.source_cutoff is None


def test_reader_degrades_invalid_quarterly_workbook_to_unavailable() -> None:
    def fetch(url: str) -> bytes:
        if url == SUGEFCapitalAdequacyReader.SOURCE_PAGE:
            return _page()
        return b"<html>not an xlsx</html>"

    reader = SUGEFCapitalAdequacyReader(
        SUGEFFinancialSourceConfig(),
        api_client=_EntityApi(),  # type: ignore[arg-type]
        fetch_bytes=fetch,
    )

    result = reader.read(date(2026, 7, 31))

    assert result.lines == ()
    assert result.source_cutoff is None
    assert result.source_files == ()
    assert any("BadZipFile" in item for item in result.diagnostics)
