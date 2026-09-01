from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException


@dataclass(frozen=True, slots=True)
class InstitutionalICLReadResult:
    source_file: str
    valuation_date: date
    sheet_name: str
    icl_total: Decimal
    icl_mn: Decimal
    icl_me: Decimal
    liquid_asset_fund_total: Decimal
    liquid_asset_fund_mn: Decimal
    liquid_asset_fund_me: Decimal
    total_outflows_30d_total: Decimal
    total_outflows_30d_mn: Decimal
    total_outflows_30d_me: Decimal
    total_inflows_30d_total: Decimal
    total_inflows_30d_mn: Decimal
    total_inflows_30d_me: Decimal
    net_cash_outflow_30d_total: Decimal
    net_cash_outflow_30d_mn: Decimal
    net_cash_outflow_30d_me: Decimal
    diagnostics: dict[str, Any] = field(default_factory=dict)
    warnings: tuple[str, ...] = ()


class InstitutionalICLReader:
    _CODE_ICL = 100000
    _CODE_LIQUID_ASSET_FUND = 200000
    _CODE_NET_OUTFLOW = 300000
    _CODE_TOTAL_OUTFLOWS = 310000
    _CODE_TOTAL_INFLOWS = 320000

    def read(self, file_path: str | Path) -> InstitutionalICLReadResult:
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(path)
        if path.suffix.lower() != ".xlsx":
            raise ValueError(f"Unsupported ICL file extension: {path.suffix}")

        try:
            workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
        except (InvalidFileException, OSError, ValueError) as exc:
            raise ValueError(f"Unable to read ICL workbook: {path.name}") from exc

        if not workbook.sheetnames:
            raise ValueError("ICL workbook contains no worksheets")

        sheet_name = workbook.sheetnames[0]
        worksheet = workbook[sheet_name]
        valuation_date = self._coerce_date(worksheet["B7"].value)
        if valuation_date is None:
            raise ValueError("ICL valuation date was not found in B7")

        rows_by_code, values_by_code = self._read_required_rows(worksheet)
        required_codes = (
            self._CODE_ICL,
            self._CODE_LIQUID_ASSET_FUND,
            self._CODE_NET_OUTFLOW,
            self._CODE_TOTAL_OUTFLOWS,
            self._CODE_TOTAL_INFLOWS,
        )
        missing_codes = [code for code in required_codes if code not in rows_by_code]
        if missing_codes:
            raise ValueError(f"Required ICL regulatory codes not found: {missing_codes}")

        icl = values_by_code[self._CODE_ICL]
        liquid_asset_fund = values_by_code[self._CODE_LIQUID_ASSET_FUND]
        net_outflow = values_by_code[self._CODE_NET_OUTFLOW]
        total_outflows = values_by_code[self._CODE_TOTAL_OUTFLOWS]
        total_inflows = values_by_code[self._CODE_TOTAL_INFLOWS]

        warnings: list[str] = []
        self._validate_difference(total_outflows, total_inflows, net_outflow, warnings)
        self._validate_icl_ratio(liquid_asset_fund, net_outflow, icl, warnings)

        diagnostics = {
            "source_cells": {
                "valuation_date": "B7",
                "icl": self._trace_cells(rows_by_code[self._CODE_ICL]),
                "liquid_asset_fund": self._trace_cells(rows_by_code[self._CODE_LIQUID_ASSET_FUND]),
                "net_cash_outflow_30d": self._trace_cells(rows_by_code[self._CODE_NET_OUTFLOW]),
                "total_outflows_30d": self._trace_cells(rows_by_code[self._CODE_TOTAL_OUTFLOWS]),
                "total_inflows_30d": self._trace_cells(rows_by_code[self._CODE_TOTAL_INFLOWS]),
            },
        }

        return InstitutionalICLReadResult(
            source_file=path.name,
            valuation_date=valuation_date,
            sheet_name=sheet_name,
            icl_total=icl[0],
            icl_mn=icl[1],
            icl_me=icl[2],
            liquid_asset_fund_total=liquid_asset_fund[0],
            liquid_asset_fund_mn=liquid_asset_fund[1],
            liquid_asset_fund_me=liquid_asset_fund[2],
            total_outflows_30d_total=total_outflows[0],
            total_outflows_30d_mn=total_outflows[1],
            total_outflows_30d_me=total_outflows[2],
            total_inflows_30d_total=total_inflows[0],
            total_inflows_30d_mn=total_inflows[1],
            total_inflows_30d_me=total_inflows[2],
            net_cash_outflow_30d_total=net_outflow[0],
            net_cash_outflow_30d_mn=net_outflow[1],
            net_cash_outflow_30d_me=net_outflow[2],
            diagnostics=diagnostics,
            warnings=tuple(warnings),
        )

    def _read_required_rows(
        self, worksheet: Any
    ) -> tuple[dict[int, int], dict[int, tuple[Decimal, Decimal, Decimal]]]:
        required_codes = {
            self._CODE_ICL,
            self._CODE_LIQUID_ASSET_FUND,
            self._CODE_NET_OUTFLOW,
            self._CODE_TOTAL_OUTFLOWS,
            self._CODE_TOTAL_INFLOWS,
        }
        rows_by_code: dict[int, int] = {}
        values_by_code: dict[int, tuple[Decimal, Decimal, Decimal]] = {}

        for row_number, values in enumerate(
            worksheet.iter_rows(
                min_row=1,
                max_row=worksheet.max_row,
                min_col=11,
                max_col=29,
                values_only=True,
            ),
            start=1,
        ):
            code = self._coerce_int(values[0])
            if code is None or code not in required_codes or code in rows_by_code:
                continue

            rows_by_code[code] = row_number
            values_by_code[code] = (
                self._decimal(values[16]),
                self._decimal(values[17]),
                self._decimal(values[18]),
            )

            if len(rows_by_code) == len(required_codes):
                break

        return rows_by_code, values_by_code

    def _validate_difference(self, total_outflows, total_inflows, net_outflow, warnings) -> None:
        labels = ("TOTAL", "MN", "ME")
        for index, label in enumerate(labels):
            expected = total_outflows[index] - total_inflows[index]
            if abs(expected - net_outflow[index]) > Decimal("1"):
                warnings.append(
                    f"Net outflow reconciliation difference for {label}: expected={expected} reported={net_outflow[index]}"
                )

    def _validate_icl_ratio(self, liquid_asset_fund, net_outflow, icl, warnings) -> None:
        labels = ("TOTAL", "MN", "ME")
        for index, label in enumerate(labels):
            denominator = net_outflow[index]
            if denominator == 0:
                warnings.append(f"ICL denominator is zero for {label}")
                continue
            expected = liquid_asset_fund[index] / denominator
            if abs(expected - icl[index]) > Decimal("0.0001"):
                warnings.append(
                    f"ICL reconciliation difference for {label}: expected={expected} reported={icl[index]}"
                )

    @staticmethod
    def _trace_cells(row_number: int) -> dict[str, str]:
        return {
            "code": f"K{row_number}",
            "consolidated": f"AA{row_number}",
            "mn": f"AB{row_number}",
            "me": f"AC{row_number}",
        }

    @staticmethod
    def _decimal(value: Any) -> Decimal:
        if value in (None, ""):
            return Decimal("0")
        try:
            return Decimal(str(value))
        except (InvalidOperation, ValueError, TypeError) as exc:
            raise ValueError(f"Invalid numeric ICL value: {value!r}") from exc

    @staticmethod
    def _coerce_int(value: Any) -> int | None:
        if value in (None, "") or isinstance(value, bool):
            return None
        try:
            return int(Decimal(str(value)))
        except (InvalidOperation, ValueError, TypeError):
            return None

    @staticmethod
    def _coerce_date(value: Any) -> date | None:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, str) and value:
            try:
                return date.fromisoformat(value)
            except ValueError:
                return None
        return None
