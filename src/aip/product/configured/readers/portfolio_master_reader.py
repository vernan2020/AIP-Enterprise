from __future__ import annotations

import math
import re
import unicodedata
import zipfile
from dataclasses import dataclass, field
from datetime import date, datetime, time, timedelta
from decimal import Decimal
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

try:
    import xlrd
except ImportError:  # pragma: no cover - exercised when xlrd is unavailable
    xlrd = None
    XLRDError = ValueError
else:
    XLRDError = xlrd.biffh.XLRDError


@dataclass(frozen=True, slots=True)
class PortfolioMasterReadResult:
    source_file: str
    valuation_date: date
    sheet_selected: str
    normalized_positions: list[dict[str, Any]]
    warnings: tuple[str, ...]
    rejected_row_count: int
    source_status: str
    detected_column_mapping: dict[str, str]
    diagnostics: dict[str, Any] = field(default_factory=dict)


class PortfolioMasterReader:
    _CANONICAL_ALIASES: dict[str, tuple[str, ...]] = {
        "isin": ("isin", "isin code", "identifier", "instrument identifier"),
        "issuer": ("issuer", "emisor", "issuer name", "nombre emisor"),
        "instrument": ("instrument", "instrumento", "security", "nombre instrumento"),
        "currency": ("currency", "moneda", "ccy", "curr"),
        "nominal_value": (
            "nominal",
            "nominal value",
            "nominal valor",
            "nominales",
            "saldo principal",
        ),
        "market_value": (
            "market value",
            "valor de mercado",
            "valor mercado",
            "mercado",
            "saldo valor mercado",
            "valor mercado colonizado",
        ),
        "book_value": (
            "book value",
            "valor en libros",
            "valor libro",
            "libros",
            "saldo valor compra",
        ),
        "purchase_value": (
            "purchase value",
            "valor compra",
            "valor de compra",
            "saldo valor transado",
        ),
        "market_price": ("market price", "precio mercado", "precio de mercado"),
        "acquisition_price": ("acquisition price", "precio adquisicion", "precio de adquisición"),
        "yield_value": (
            "yield",
            "rendimiento",
            "yield value",
            "market yield",
            "rendimiento mercado",
            "tir",
        ),
        "nominal_rate": ("nominal rate", "tasa nominal", "rate"),
        "modified_duration": (
            "modified duration",
            "duracion modificada",
            "duración modificada",
            "duration",
            "duracion",
        ),
        "maturity_date": (
            "maturity date",
            "fecha vencimiento",
            "fecha de vencimiento",
            "vencimiento",
        ),
        "next_coupon_date": (
            "next coupon date",
            "proxima cupon",
            "próxima cupón",
            "fecha proximo cupon",
        ),
        "rate_type": ("rate type", "tipo tasa", "tipo de tasa"),
        "classification": ("classification", "clasificacion", "clasificación", "category"),
        "reserve_liquidity_indicator": (
            "reserve liquidity",
            "liquidity",
            "liquidez",
            "indicador liquidez",
        ),
        "encumbrance_status": ("encumbrance", "encumbrado", "commitment", "compromiso"),
        "accounting_account": ("account", "accounting account", "cuenta", "cuenta contable"),
        "source_cutoff": ("source cutoff", "cutoff", "fecha corte", "corte"),
    }

    def read(
        self, path: str | Path, *, valuation_date_override: date | None = None
    ) -> PortfolioMasterReadResult:
        file_path = Path(path)
        if not file_path.exists():
            return PortfolioMasterReadResult(
                source_file=str(file_path),
                valuation_date=valuation_date_override or date.today(),
                sheet_selected="",
                normalized_positions=[],
                warnings=("Portfolio master file was not found",),
                rejected_row_count=0,
                source_status="UNAVAILABLE",
                detected_column_mapping={},
                diagnostics={"error": "file_not_found"},
            )

        valuation_date = valuation_date_override or self._infer_valuation_date(file_path)
        warnings: list[str] = []
        rejected_rows = 0
        positions: list[dict[str, Any]] = []
        seen_identities: set[str] = set()

        workbook_type = self._detect_workbook_type(file_path)
        try:
            if workbook_type == "xls":
                if xlrd is None:
                    raise RuntimeError("xlrd is required to read .xls files")
                workbook = xlrd.open_workbook(file_path)
                sheets = workbook.sheets()
                sheet_data = [self._read_xls_sheet(sheet) for sheet in sheets]
                sheet_names = [sheet.name for sheet in sheets]
            else:
                workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                sheet_names = workbook.sheetnames
                sheet_data = [self._read_xlsx_sheet(sheet) for sheet in workbook.worksheets]
                workbook.close()
        except (
            InvalidFileException,
            XLRDError,
            zipfile.BadZipFile,
            ValueError,
            OSError,
            RuntimeError,
        ) as exc:
            return PortfolioMasterReadResult(
                source_file=str(file_path),
                valuation_date=valuation_date,
                sheet_selected="",
                normalized_positions=[],
                warnings=(f"Workbook could not be read: {exc}",),
                rejected_row_count=0,
                source_status="UNAVAILABLE",
                detected_column_mapping={},
                diagnostics={"workbook_type": workbook_type, "error": str(exc), "sheet_count": 0},
            )

        sheet_index, header_index, header_values, normalized_headers, column_map = (
            self._select_sheet(sheet_names, sheet_data)
        )
        if sheet_index is None or header_index is None:
            return PortfolioMasterReadResult(
                source_file=str(file_path),
                valuation_date=valuation_date,
                sheet_selected="",
                normalized_positions=[],
                warnings=("No institutional header row could be detected",),
                rejected_row_count=0,
                source_status="UNAVAILABLE",
                detected_column_mapping={},
                diagnostics={
                    "workbook_type": workbook_type,
                    "sheet_count": len(sheet_names),
                    "sheets": sheet_names,
                },
            )

        sheet_name = sheet_names[sheet_index]
        rows = sheet_data[sheet_index]
        for row_index, row_values in enumerate(rows[header_index + 1 :], start=header_index + 2):
            if self._is_blank_row(row_values):
                continue

            normalized_row = self._normalize_row(row_values, header_values)
            if normalized_row is None:
                rejected_rows += 1
                warnings.append(f"Rejected row {row_index}: blank row")
                continue

            position = self._build_position(normalized_row, row_index, file_path.name)
            if not position:
                rejected_rows += 1
                warnings.append(f"Rejected row {row_index}: malformed position")
                continue

            position_identity = self._build_position_identity(position)
            if position_identity in seen_identities:
                rejected_rows += 1
                warnings.append(
                    f"Duplicate identity detected for row {row_index}: {position_identity}"
                )
                continue

            seen_identities.add(position_identity)
            positions.append(position)

        missing_columns = [name for name in ("isin", "market_value") if name not in column_map]
        if missing_columns:
            warnings.append(f"Missing expected columns: {', '.join(missing_columns)}")

        source_status = "HEALTHY"
        if warnings or rejected_rows or len(positions) == 0:
            source_status = "DEGRADED" if positions else "UNAVAILABLE"
        diagnostics = {
            "workbook_type": workbook_type,
            "sheet_count": len(sheet_names),
            "sheet": sheet_name,
            "header_row": header_index + 1,
            "rows_read": len(rows),
            "rows_accepted": len(positions),
            "rows_rejected": rejected_rows,
            "column_mapping": column_map,
            "normalized_headers": normalized_headers,
        }
        return PortfolioMasterReadResult(
            source_file=str(file_path),
            valuation_date=valuation_date,
            sheet_selected=sheet_name,
            normalized_positions=positions,
            warnings=tuple(warnings),
            rejected_row_count=rejected_rows,
            source_status=source_status,
            detected_column_mapping=column_map,
            diagnostics=diagnostics,
        )

    def _detect_workbook_type(self, path: Path) -> str:
        return "xls" if path.suffix.lower() == ".xls" else "xlsx"

    def _read_xls_sheet(self, sheet: xlrd.sheet.Sheet) -> list[list[Any]]:
        rows: list[list[Any]] = []
        for row_index in range(sheet.nrows):
            values = []
            for cell_index in range(sheet.ncols):
                cell = sheet.cell(row_index, cell_index)
                if cell.ctype == xlrd.XL_CELL_DATE:
                    values.append(xlrd.xldate_as_datetime(cell.value, sheet.book.datemode))
                elif cell.ctype == xlrd.XL_CELL_NUMBER:
                    values.append(cell.value)
                elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                    values.append(bool(cell.value))
                elif cell.ctype == xlrd.XL_CELL_ERROR:
                    values.append("")
                else:
                    values.append(cell.value)
            rows.append(values)
        return rows

    def _read_xlsx_sheet(self, sheet: openpyxl.worksheet.worksheet.Worksheet) -> list[list[Any]]:
        rows: list[list[Any]] = []
        for row_values in sheet.iter_rows(values_only=True):
            rows.append(list(row_values))
        return rows

    def _select_sheet(
        self, sheet_names: list[str], sheet_data: list[list[list[Any]]]
    ) -> tuple[int | None, int | None, list[str], list[str], dict[str, str]]:
        best_score = -1
        best_selection: (
            tuple[int | None, int | None, list[str], list[str], dict[str, str]] | None
        ) = None
        for sheet_index, rows in enumerate(sheet_data):
            for header_index, header_row in enumerate(rows[: min(len(rows), 20)]):
                normalized_headers = [
                    self._normalize_header(self._stringify(value)) for value in header_row
                ]
                if not normalized_headers:
                    continue
                matches = self._match_header_row(normalized_headers)
                score = sum(1 for _, _ in matches)
                if score < 3:
                    continue
                column_map: dict[str, str] = {}
                for canonical, aliases in self._CANONICAL_ALIASES.items():
                    for index, header in enumerate(normalized_headers):
                        if self._matches_alias(header, aliases):
                            column_map[canonical] = self._stringify(header_row[index])
                            break
                if score >= 5:
                    best_score = score
                    best_selection = (
                        sheet_index,
                        header_index,
                        [self._stringify(value) for value in header_row],
                        normalized_headers,
                        column_map,
                    )
                    break
                for canonical, aliases in self._CANONICAL_ALIASES.items():
                    for index, header in enumerate(normalized_headers):
                        if self._matches_alias(header, aliases):
                            column_map[canonical] = self._stringify(header_row[index])
                            break
                if score > best_score:
                    best_score = score
                    best_selection = (
                        sheet_index,
                        header_index,
                        [self._stringify(value) for value in header_row],
                        normalized_headers,
                        column_map,
                    )
        if best_selection is None:
            return None, None, [], [], {}
        return best_selection

    def _match_header_row(self, normalized_headers: list[str]) -> list[tuple[str, str]]:
        matches: list[tuple[str, str]] = []
        for canonical, aliases in self._CANONICAL_ALIASES.items():
            for header in normalized_headers:
                if self._matches_alias(header, aliases):
                    matches.append((canonical, header))
                    break
        return matches

    def _normalize_header(self, value: str) -> str:
        text = str(value).strip().lower()
        normalized = unicodedata.normalize("NFKD", text)
        ascii_text = normalized.encode("ascii", "ignore").decode("ascii")
        ascii_text = re.sub(r"\s+", " ", ascii_text)
        return re.sub(r"[^a-z0-9]+", " ", ascii_text).strip()

    def _matches_alias(self, header: str, aliases: tuple[str, ...]) -> bool:
        normalized_header = self._normalize_header(header)
        if not normalized_header:
            return False
        for alias in aliases:
            alias_normalized = self._normalize_header(alias)
            if not alias_normalized:
                continue
            if (
                normalized_header == alias_normalized
                or normalized_header.startswith(alias_normalized)
                or alias_normalized.startswith(normalized_header)
            ):
                return True
        return False

    def _stringify(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, (datetime, date, time)):
            return value.isoformat()
        if isinstance(value, Decimal):
            return format(value, "f")
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            return str(value)
        return str(value)

    def _normalize_row(
        self, row_values: list[Any], header_values: list[str]
    ) -> dict[str, Any] | None:
        values = [self._stringify(value) for value in row_values]
        if not any(value.strip() for value in values):
            return None
        result: dict[str, Any] = {}
        for index, header in enumerate(header_values):
            if index >= len(row_values):
                continue
            result[self._normalize_header(header)] = row_values[index]
        return result

    def _build_position(
        self, row_data: dict[str, Any], row_number: int, file_name: str
    ) -> dict[str, Any] | None:
        normalized_cells = {self._normalize_header(key): value for key, value in row_data.items()}
        isin = self._coerce_text(self._find_field(normalized_cells, "isin"))
        issuer = self._coerce_text(self._find_field(normalized_cells, "issuer"))
        instrument = self._coerce_text(self._find_field(normalized_cells, "instrument"))
        currency = self._normalize_currency(
            self._coerce_text(self._find_field(normalized_cells, "currency"))
        )
        nominal_value = self._parse_number(self._find_field(normalized_cells, "nominal_value"))
        market_value = self._parse_number(self._find_field(normalized_cells, "market_value"))
        book_value = self._parse_number(self._find_field(normalized_cells, "book_value"))
        yield_value = self._parse_number(
            self._find_field(normalized_cells, "yield_value"), percentage=True
        )
        modified_duration = self._parse_number(
            self._find_field(normalized_cells, "modified_duration")
        )
        maturity_date = self._parse_date(self._find_field(normalized_cells, "maturity_date"))
        classification = self._coerce_text(self._find_field(normalized_cells, "classification"))
        account = self._coerce_text(self._find_field(normalized_cells, "accounting_account"))
        if (
            not isin
            and not issuer
            and not instrument
            and not market_value
            and not book_value
            and not nominal_value
        ):
            return None
        if not isin and not issuer and not instrument:
            return None
        if market_value is None and book_value is None and nominal_value is None:
            return None
        if not isin:
            isin = f"{file_name}:{row_number}"
        return {
            "source_file": file_name,
            "source_row": row_number,
            "source_values": {key: self._stringify(value) for key, value in row_data.items()},
            "isin": isin,
            "issuer": issuer,
            "instrument": instrument or "Instrument",
            "currency": currency or "USD",
            "nominal": float(nominal_value) if nominal_value is not None else 0.0,
            "market_value": float(market_value) if market_value is not None else 0.0,
            "book_value": float(book_value) if book_value is not None else 0.0,
            "yield_value": float(yield_value) if yield_value is not None else 0.0,
            "modified_duration": float(modified_duration) if modified_duration is not None else 0.0,
            "classification": classification or "Unknown",
            "hqla_status": "Unknown",
            "mil_status": "Unknown",
            "recommendation": "Hold",
            "encumbered": False,
            "account": account,
            "maturity_date": maturity_date.isoformat() if maturity_date else None,
            "source_cutoff": None,
        }

    def _find_field(self, normalized_cells: dict[str, Any], canonical: str) -> Any:
        aliases = self._CANONICAL_ALIASES.get(canonical, ())
        for alias in aliases:
            alias_key = self._normalize_header(alias)
            if alias_key in normalized_cells:
                return normalized_cells[alias_key]
        return None

    def _coerce_text(self, value: Any) -> str:
        if value is None:
            return ""
        text = self._stringify(value).strip()
        return re.sub(r"\s+", " ", text)

    def _normalize_currency(self, value: str) -> str:
        cleaned = self._coerce_text(value).upper()
        if cleaned in {"CRC", "COLONES", "CRC"}:
            return "CRC"
        if cleaned in {"USD", "US$", "$"}:
            return "USD"
        return cleaned

    def _parse_number(self, value: Any, *, percentage: bool = False) -> Decimal | None:
        if value is None:
            return None
        if isinstance(value, bool):
            return None
        if isinstance(value, (int, float, Decimal)):
            numeric_value = Decimal(str(value))
        else:
            text = self._stringify(value).strip()
            if not text:
                return None
            text = text.replace("$", "").replace("€", "").replace("£", "")
            text = (
                text.replace(".", "").replace(",", ".")
                if text.count(",") == 1 and text.count(".") == 0
                else text
            )
            if "," in text and "." in text:
                if text.rfind(",") > text.rfind("."):
                    text = text.replace(".", "").replace(",", ".")
                else:
                    text = text.replace(",", "")
            elif "," in text:
                text = text.replace(",", ".")
            text = text.replace("%", "")
            if not text:
                return None
            numeric_value = Decimal(text)
        if percentage and numeric_value != 0:
            numeric_value = numeric_value / Decimal("100")
        return numeric_value

    def _parse_date(self, value: Any) -> date | None:
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            if math.isfinite(float(value)):
                try:
                    return datetime(1899, 12, 30) + timedelta(days=int(value))
                except ValueError:
                    return None
        text = self._stringify(value).strip()
        if not text:
            return None
        try:
            return date.fromisoformat(text)
        except ValueError:
            for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%Y/%m/%d"):
                try:
                    return datetime.strptime(text, fmt).date()
                except ValueError:
                    continue
        return None

    def _build_position_identity(self, position: dict[str, Any]) -> str:
        parts = [
            self._coerce_text(position.get("isin")),
            self._coerce_text(position.get("account")),
            self._coerce_text(position.get("issuer")),
            self._coerce_text(position.get("maturity_date", "")),
        ]
        return "|".join(parts)

    def _infer_valuation_date(self, file_path: Path) -> date:
        for candidate in (file_path.name, file_path.stem):
            parsed = self._parse_date_from_name(candidate)
            if parsed is not None:
                return parsed
        return date.today()

    def _parse_date_from_name(self, value: str) -> date | None:
        compact_match = re.search(r"(?P<year>\d{4})(?P<month>\d{2})(?P<day>\d{2})", value)
        if compact_match is not None:
            try:
                return date(
                    int(compact_match.group("year")),
                    int(compact_match.group("month")),
                    int(compact_match.group("day")),
                )
            except ValueError:
                return None
        date_match = re.search(
            r"(?P<day>\d{1,2})[-.]?(?P<month>\d{1,2})[-.]?(?P<year>\d{4})", value
        )
        if not date_match:
            return None
        try:
            return date(
                int(date_match.group("year")),
                int(date_match.group("month")),
                int(date_match.group("day")),
            )
        except ValueError:
            return None

    def _is_blank_row(self, row_values: list[Any]) -> bool:
        if not row_values:
            return True
        return all(self._stringify(value).strip() == "" for value in row_values)
