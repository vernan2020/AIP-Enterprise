from __future__ import annotations

import math
import re
import unicodedata
import zipfile
from dataclasses import dataclass, field
from datetime import date, datetime, time, timedelta
from decimal import Decimal
from pathlib import Path
from typing import Any

import openpyxl
import xlrd
from openpyxl.utils.exceptions import InvalidFileException


@dataclass(frozen=True, slots=True)
class InstitutionalPortfolioMasterReadResult:
    source_file: str
    valuation_date: date
    sheet_selected: str
    normalized_positions: list[dict[str, Any]]
    warnings: tuple[str, ...]
    rejected_row_count: int
    source_status: str
    detected_column_mapping: dict[str, str]
    diagnostics: dict[str, Any] = field(default_factory=dict)


class InstitutionalPortfolioMasterReader:
    _CANONICAL_ALIASES: dict[str, tuple[str, ...]] = {
        "source_row": ("#", "source row", "source_row", "row"),
        "issuer": ("emisor", "issuer", "issuer name", "nombre emisor"),
        "contract_number": ("numero contrato", "numero de contrato", "numero contrato", "número contrato", "contract number"),
        "broker": ("puesto bolsa", "broker", "brokerage", "puesto"),
        "currency": ("moneda", "currency", "ccy"),
        "acquisition_date": ("fecha ingreso", "fecha de ingreso", "acquisition date"),
        "issue_date": ("fecha emision", "fecha de emision", "issue date"),
        "maturity_date": ("fecha vencimiento", "fecha de vencimiento", "maturity date"),
        "days_to_maturity": ("dias al vencimiento", "días al vencimiento", "days to maturity"),
        "product_code": ("codigo producto", "código producto", "product code"),
        "classification": ("clasificacion", "clasificación", "classification"),
        "risk_rating": ("calificacion riesgo", "calificación riesgo", "risk rating"),
        "liquidity_reserve_flag": ("reserva liquidez", "reserve liquidity"),
        "series": ("serie", "series", "series number"),
        "market_value_crc": ("valor mercado colonizado", "market value colonized"),
        "isin": ("isin", "isin code", "instrument identifier", "identifier"),
        "traded_balance": ("saldo valor transado", "traded balance"),
        "principal_balance": ("saldo principal", "principal balance"),
        "purchase_price_percentage": ("porcentaje valor compra", "purchase price percentage"),
        "book_value": ("saldo valor compra", "book value", "valor en libros", "valor libro"),
        "market_price_percentage": ("porcentaje valor mercado", "market price percentage"),
        "market_value": ("saldo valor mercado", "market value", "valor de mercado", "valor mercado"),
        "nominal_rate": ("tasa nominal", "nominal rate"),
        "periodicity": ("periodicidad", "periodicity"),
        "accrued_interest": ("interes por cobrar", "interés por cobrar", "accrued interest"),
        "last_interest_payment_date": ("fecha ultimo pago intereses", "fecha último pago intereses", "last interest payment date"),
        "participation_quantity": ("cantidad participaciones", "participation quantity"),
        "portfolio_yield": ("tir", "portfolio yield", "yield"),
        "variable_rate_flag": ("indicador tasa variable", "indicator rate variable"),
        "public_private_indicator": ("indicador pulblico", "indicador publico", "indicador público", "public private indicator"),
        "estimated_loss": ("monto estimacion", "monto estimación", "estimated loss"),
        "impairment_amount": ("monto deterioro", "impairment amount"),
        "custodian": ("custodio", "custodian"),
    }

    _REQUIRED_COLUMNS = ("isin", "market_value", "book_value")
    _MAX_DIAGNOSTIC_TRACE_ENTRIES = 20
    _MAX_ACCEPTED_FIELD_DIAGNOSTICS = 3

    def read(self, path: str | Path, *, valuation_date_override: date | None = None, diagnostic_mode: bool = False) -> InstitutionalPortfolioMasterReadResult:
        file_path = Path(path)
        if not file_path.exists():
            return InstitutionalPortfolioMasterReadResult(
                source_file=str(file_path),
                valuation_date=valuation_date_override or date.today(),
                sheet_selected="",
                normalized_positions=[],
                warnings=("Portfolio master file was not found",),
                rejected_row_count=0,
                source_status="UNAVAILABLE",
                detected_column_mapping={},
                diagnostics={"error": "file_not_found"},
            )

        valuation_date = valuation_date_override or self._infer_valuation_date(file_path)
        workbook_type = self._detect_workbook_type(file_path)
        try:
            if workbook_type == "xls":
                workbook = xlrd.open_workbook(file_path)
                sheet_names = workbook.sheet_names()
                sheet_rows = [self._read_xls_sheet(sheet) for sheet in workbook.sheets()]
            else:
                workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                sheet_names = workbook.sheetnames
                sheet_rows = [self._read_xlsx_sheet(sheet) for sheet in workbook.worksheets]
                workbook.close()
        except (InvalidFileException, xlrd.biffh.XLRDError, zipfile.BadZipFile, ValueError, OSError) as exc:
            return InstitutionalPortfolioMasterReadResult(
                source_file=str(file_path),
                valuation_date=valuation_date,
                sheet_selected="",
                normalized_positions=[],
                warnings=(f"Workbook could not be read: {exc}",),
                rejected_row_count=0,
                source_status="UNAVAILABLE",
                detected_column_mapping={},
                diagnostics={"workbook_type": workbook_type, "error": str(exc), "sheet_count": 0},
            )

        sheet_index, header_index, header_values, normalized_headers, column_map, column_indices = self._select_sheet(sheet_names, sheet_rows)
        if sheet_index is None or header_index is None:
            return InstitutionalPortfolioMasterReadResult(
                source_file=str(file_path),
                valuation_date=valuation_date,
                sheet_selected="",
                normalized_positions=[],
                warnings=("No institutional header row could be detected",),
                rejected_row_count=0,
                source_status="UNAVAILABLE",
                detected_column_mapping={},
                diagnostics={"workbook_type": workbook_type, "sheet_count": len(sheet_names), "sheets": sheet_names},
            )

        rows = sheet_rows[sheet_index]
        positions: list[dict[str, Any]] = []
        warnings: list[str] = []
        rejected_rows = 0
        accepted_field_diagnostics: list[dict[str, Any]] = []
        identities: set[str] = set()
        duplicate_identities: list[str] = []
        record_trace: list[dict[str, Any]] = []
        first_rejected_row: dict[str, Any] | None = None
        for row_index, row_values in enumerate(rows[header_index + 1 :], start=header_index + 2):
            if self._is_blank_row(row_values):
                rejected_rows += 1
                if diagnostic_mode and first_rejected_row is None:
                    first_rejected_row = self._build_rejection_diagnostic(row_values, row_index, "blank_row", normalized_row=None)
                if diagnostic_mode:
                    record_trace.append({"row": row_index, "status": "discarded", "reason": "blank_row", "isin": None})
                continue
            normalized_row = self._normalize_row(row_values, header_values)
            if normalized_row is None:
                rejected_rows += 1
                warnings.append(f"Rejected row {row_index}: blank row")
                if diagnostic_mode and first_rejected_row is None:
                    first_rejected_row = self._build_rejection_diagnostic(row_values, row_index, "blank_row", normalized_row=None)
                if diagnostic_mode:
                    record_trace.append({"row": row_index, "status": "discarded", "reason": "blank_row", "isin": None})
                continue

            position = self._build_position(normalized_row, row_index, file_path.name)
            if position is None:
                rejected_rows += 1
                warnings.append(f"Rejected row {row_index}: malformed position")
                if diagnostic_mode and first_rejected_row is None:
                    first_rejected_row = self._build_rejection_diagnostic(row_values, row_index, "malformed_position", normalized_row=normalized_row, position=position)
                if diagnostic_mode:
                    record_trace.append({"row": row_index, "status": "discarded", "reason": "malformed_position", "isin": None})
                continue

            position_identity = self._build_position_identity(position)
            if position_identity in identities:
                rejected_rows += 1
                duplicate_identities.append(position_identity)
                warnings.append(f"Duplicate identity detected for row {row_index}: {position_identity}")
                if diagnostic_mode and first_rejected_row is None:
                    first_rejected_row = self._build_rejection_diagnostic(row_values, row_index, "duplicate_identity", normalized_row=normalized_row, position=position)
                if diagnostic_mode:
                    record_trace.append({"row": row_index, "status": "discarded", "reason": "duplicate_identity", "isin": position.get("isin")})
                continue

            identities.add(position_identity)
            positions.append(position)
            if diagnostic_mode:
                record_trace.append({"row": row_index, "status": "accepted", "reason": "parsed", "isin": position.get("isin")})
                if len(accepted_field_diagnostics) < self._MAX_ACCEPTED_FIELD_DIAGNOSTICS:
                    accepted_field_diagnostics.append(
                        self._build_field_diagnostics(row_values, header_values, column_indices, row_index)
                    )

        missing_columns = [column for column in self._REQUIRED_COLUMNS if column not in column_map]
        if missing_columns:
            warnings.append(f"Missing mandatory columns: {', '.join(missing_columns)}")

        source_status = "HEALTHY"
        if warnings or rejected_rows or len(positions) == 0:
            source_status = "DEGRADED" if positions else "UNAVAILABLE"

        trace_payload = None
        if diagnostic_mode:
            trace_payload = {
                "source_file": self._safe_diagnostic_reference(file_path),
                "sheet_used": sheet_names[sheet_index],
                "header_row": header_index + 1,
                "records_read": len(rows[header_index + 1 :]),
                "records_valid": len(positions),
                "records_discarded": rejected_rows,
                "isin_found": [position.get("isin") for position in positions if position.get("isin")],
                "record_trace": record_trace[-self._MAX_DIAGNOSTIC_TRACE_ENTRIES :],
                "first_accepted_row_field_diagnostics": accepted_field_diagnostics,
            }

        diagnostics = {
            "workbook_type": workbook_type,
            "sheet_count": len(sheet_names),
            "sheet": sheet_names[sheet_index],
            "header_row": header_index + 1,
            "rows_read": len(rows),
            "rows_accepted": len(positions),
            "rows_rejected": rejected_rows,
            "column_mapping": column_map,
            "column_indices": column_indices,
            "normalized_headers": normalized_headers,
            "missing_columns": missing_columns,
            "duplicate_identities": duplicate_identities,
            "source_reference": self._safe_diagnostic_reference(file_path),
        }
        if first_rejected_row is not None:
            diagnostics["first_rejected_row"] = first_rejected_row
        if trace_payload is not None:
            diagnostics["trace"] = trace_payload
            if first_rejected_row is not None:
                trace_payload["first_rejected_row"] = first_rejected_row
        return InstitutionalPortfolioMasterReadResult(
            source_file=str(file_path),
            valuation_date=valuation_date,
            sheet_selected=sheet_names[sheet_index],
            normalized_positions=positions,
            warnings=tuple(warnings),
            rejected_row_count=rejected_rows,
            source_status=source_status,
            detected_column_mapping=column_map,
            diagnostics=diagnostics,
        )

    def _detect_workbook_type(self, path: Path) -> str:
        return "xls" if path.suffix.lower() == ".xls" else "xlsx"

    def _safe_diagnostic_reference(self, path: Path | str | None) -> str:
        if not path:
            return ""
        path_text = str(path)
        if not path_text:
            return ""
        candidate = Path(path_text)
        if not candidate.name:
            return ""
        return candidate.name

    def _read_xls_sheet(self, sheet: xlrd.sheet.Sheet) -> list[list[Any]]:
        rows: list[list[Any]] = []
        for row_index in range(sheet.nrows):
            values: list[Any] = []
            for cell_index in range(sheet.ncols):
                cell = sheet.cell(row_index, cell_index)
                if cell.ctype == xlrd.XL_CELL_DATE:
                    values.append(xlrd.xldate_as_datetime(cell.value, sheet.book.datemode))
                elif cell.ctype == xlrd.XL_CELL_NUMBER:
                    values.append(cell.value)
                elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                    values.append(bool(cell.value))
                elif cell.ctype == xlrd.XL_CELL_ERROR:
                    values.append("")
                else:
                    values.append(cell.value)
            rows.append(values)
        return rows

    def _read_xlsx_sheet(self, sheet: openpyxl.worksheet.worksheet.Worksheet) -> list[list[Any]]:
        rows: list[list[Any]] = []
        for row_values in sheet.iter_rows(values_only=True):
            rows.append(list(row_values))
        return rows

    def _select_sheet(self, sheet_names: list[str], sheet_rows: list[list[list[Any]]]) -> tuple[int | None, int | None, list[str], list[str], dict[str, str], dict[str, int]]:
        best_score = -1
        best_selection: tuple[int | None, int | None, list[str], list[str], dict[str, str], dict[str, int]] | None = None
        for sheet_index, rows in enumerate(sheet_rows):
            for header_index, header_row in enumerate(rows[: min(len(rows), 20)]):
                normalized_headers = [self._normalize_header(self._stringify(value)) for value in header_row]
                if not normalized_headers:
                    continue
                matches = self._match_header_row(normalized_headers)
                score = sum(1 for _, _ in matches)
                if score < 3:
                    continue
                column_map, column_indices = self._build_column_mapping(header_row, normalized_headers)
                if score >= 5:
                    return sheet_index, header_index, [self._stringify(value) for value in header_row], normalized_headers, column_map, column_indices
                if score > best_score:
                    best_score = score
                    best_selection = (sheet_index, header_index, [self._stringify(value) for value in header_row], normalized_headers, column_map, column_indices)
        if best_selection is None:
            return None, None, [], [], {}, {}
        return best_selection

    def _match_header_row(self, normalized_headers: list[str]) -> list[tuple[str, str]]:
        matches: list[tuple[str, str]] = []
        for canonical, aliases in self._CANONICAL_ALIASES.items():
            for header in normalized_headers:
                if self._matches_alias(header, aliases):
                    matches.append((canonical, header))
                    break
        return matches

    def _build_column_mapping(self, header_values: list[str], normalized_headers: list[str]) -> tuple[dict[str, str], dict[str, int]]:
        column_map: dict[str, str] = {}
        column_indices: dict[str, int] = {}
        for canonical, aliases in self._CANONICAL_ALIASES.items():
            for index, header in enumerate(normalized_headers):
                if self._matches_alias(header, aliases):
                    column_map[canonical] = self._stringify(header_values[index])
                    column_indices[canonical] = index
                    break
        return column_map, column_indices

    def _normalize_header(self, value: str) -> str:
        text = self._stringify(value)
        text = text.replace("\ufeff", "")
        text = text.replace("\u00a0", " ")
        text = text.replace("_", " ")
        text = re.sub(r"[\u200b\u200c\u200d\ufeff\u2060]", "", text)
        text = text.strip().lower()
        normalized = unicodedata.normalize("NFKD", text)
        ascii_text = normalized.encode("ascii", "ignore").decode("ascii")
        ascii_text = re.sub(r"\s+", " ", ascii_text)
        ascii_text = re.sub(r"[^a-z0-9]+", " ", ascii_text)
        return re.sub(r"\s+", " ", ascii_text).strip()

    def _matches_alias(self, header: str, aliases: tuple[str, ...]) -> bool:
        normalized_header = self._normalize_header(header)
        if not normalized_header:
            return False
        for alias in aliases:
            alias_normalized = self._normalize_header(alias)
            if not alias_normalized:
                continue
            if normalized_header == alias_normalized or normalized_header.startswith(alias_normalized) or alias_normalized.startswith(normalized_header):
                return True
        return False

    def _stringify(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, (datetime, date, time)):
            return value.isoformat()
        if isinstance(value, Decimal):
            return format(value, "f")
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            return str(value)
        return str(value)

    def _normalize_row(self, row_values: list[Any], header_values: list[str]) -> dict[str, Any] | None:
        values = [self._stringify(value) for value in row_values]
        if not any(value.strip() for value in values):
            return None
        result: dict[str, Any] = {}
        for index, header in enumerate(header_values):
            if index >= len(row_values):
                continue
            result[self._normalize_header(header)] = row_values[index]
        return result

    def _is_blank_row(self, row_values: list[Any]) -> bool:
        return not any(self._stringify(value).strip() for value in row_values)

    def _build_position(self, row_data: dict[str, Any], row_number: int, file_name: str) -> dict[str, Any] | None:
        normalized_cells = {self._normalize_header(key): value for key, value in row_data.items()}
        isin = self._coerce_text(self._find_field(normalized_cells, "isin"))
        issuer = self._coerce_text(self._find_field(normalized_cells, "issuer"))
        contract_number = self._coerce_text(self._find_field(normalized_cells, "contract_number"))
        broker = self._coerce_text(self._find_field(normalized_cells, "broker"))
        currency = self._normalize_currency(self._coerce_text(self._find_field(normalized_cells, "currency")))
        acquisition_date = self._parse_date(self._find_field(normalized_cells, "acquisition_date"))
        issue_date = self._parse_date(self._find_field(normalized_cells, "issue_date"))
        maturity_date = self._parse_date(self._find_field(normalized_cells, "maturity_date"))
        days_to_maturity = self._parse_number(self._find_field(normalized_cells, "days_to_maturity"))
        product_code = self._coerce_text(self._find_field(normalized_cells, "product_code"))
        classification = self._coerce_text(self._find_field(normalized_cells, "classification"))
        risk_rating = self._coerce_text(self._find_field(normalized_cells, "risk_rating"))
        liquidity_reserve_flag = self._coerce_text(self._find_field(normalized_cells, "liquidity_reserve_flag"))
        series = self._coerce_text(self._find_field(normalized_cells, "series"))
        market_value_crc = self._parse_number(self._find_field(normalized_cells, "market_value_crc"))
        traded_balance = self._parse_number(self._find_field(normalized_cells, "traded_balance"))
        principal_balance = self._parse_number(self._find_field(normalized_cells, "principal_balance"))
        purchase_price_percentage = self._parse_number(self._find_field(normalized_cells, "purchase_price_percentage"))
        book_value = self._parse_number(self._find_field(normalized_cells, "book_value"))
        market_price_percentage = self._parse_number(self._find_field(normalized_cells, "market_price_percentage"))
        market_value = self._parse_number(self._find_field(normalized_cells, "market_value"))
        nominal_rate = self._parse_number(self._find_field(normalized_cells, "nominal_rate"))
        periodicity = self._coerce_text(self._find_field(normalized_cells, "periodicity"))
        accrued_interest = self._parse_number(self._find_field(normalized_cells, "accrued_interest"))
        last_interest_payment_date = self._parse_date(self._find_field(normalized_cells, "last_interest_payment_date"))
        participation_quantity = self._parse_number(self._find_field(normalized_cells, "participation_quantity"))
        portfolio_yield = self._parse_number(self._find_field(normalized_cells, "portfolio_yield"))
        variable_rate_flag = self._coerce_text(self._find_field(normalized_cells, "variable_rate_flag"))
        public_private_indicator = self._coerce_text(self._find_field(normalized_cells, "public_private_indicator"))
        estimated_loss = self._parse_number(self._find_field(normalized_cells, "estimated_loss"))
        impairment_amount = self._parse_number(self._find_field(normalized_cells, "impairment_amount"))
        custodian = self._coerce_text(self._find_field(normalized_cells, "custodian"))

        if not any([isin, issuer, contract_number, market_value, book_value]):
            return None

        return {
            "source_row": row_number,
            "source_file": file_name,
            "issuer": issuer,
            "contract_number": contract_number,
            "broker": broker,
            "currency": currency or "CRC",
            "acquisition_date": acquisition_date,
            "issue_date": issue_date,
            "maturity_date": maturity_date,
            "days_to_maturity": float(days_to_maturity) if days_to_maturity is not None else None,
            "product_code": product_code,
            "classification": classification,
            "risk_rating": risk_rating,
            "liquidity_reserve_flag": liquidity_reserve_flag,
            "series": series,
            "market_value_crc": float(market_value_crc) if market_value_crc is not None else None,
            "isin": isin,
            "traded_balance": float(traded_balance) if traded_balance is not None else None,
            "principal_balance": float(principal_balance) if principal_balance is not None else None,
            "purchase_price_percentage": float(purchase_price_percentage) if purchase_price_percentage is not None else None,
            "book_value": float(book_value) if book_value is not None else None,
            "market_price_percentage": float(market_price_percentage) if market_price_percentage is not None else None,
            "market_value": float(market_value) if market_value is not None else None,
            "nominal_rate": float(nominal_rate) if nominal_rate is not None else None,
            "periodicity": periodicity,
            "accrued_interest": float(accrued_interest) if accrued_interest is not None else None,
            "last_interest_payment_date": last_interest_payment_date,
            "participation_quantity": float(participation_quantity) if participation_quantity is not None else None,
            "portfolio_yield": float(portfolio_yield) if portfolio_yield is not None else None,
            "variable_rate_flag": variable_rate_flag,
            "public_private_indicator": public_private_indicator,
            "estimated_loss": float(estimated_loss) if estimated_loss is not None else None,
            "impairment_amount": float(impairment_amount) if impairment_amount is not None else None,
            "custodian": custodian,
            "accounting_fields": {},
            "source_values": {key: self._stringify(value) for key, value in row_data.items()},
        }

    def _build_position_identity(self, position: dict[str, Any]) -> str:
        contract_number = str(position.get("contract_number", "")).strip()
        isin = str(position.get("isin", "")).strip()
        series = str(position.get("series", "")).strip()
        if contract_number and isin:
            return f"contract:{contract_number}|isin:{isin}"
        if contract_number and series:
            return f"contract:{contract_number}|series:{series}"
        return f"row:{position.get('source_row')}"

    def _find_field(self, normalized_cells: dict[str, Any], canonical: str) -> Any:
        aliases = self._CANONICAL_ALIASES.get(canonical, ())
        for alias in aliases:
            alias_key = self._normalize_header(alias)
            if alias_key in normalized_cells:
                return normalized_cells[alias_key]
        return None

    def _coerce_text(self, value: Any) -> str:
        if value is None:
            return ""
        text = self._stringify(value).strip()
        return re.sub(r"\s+", " ", text)

    def _build_field_diagnostics(self, row_values: list[Any], header_values: list[str], column_indices: dict[str, int], row_number: int) -> dict[str, Any]:
        diagnostics: dict[str, Any] = {"row": row_number, "fields": {}}
        for display_name, canonical in (("serie", "series"), ("codigo producto", "product_code"), ("fecha vencimiento", "maturity_date"), ("ISIN", "isin")):
            index = column_indices.get(canonical)
            if index is None:
                diagnostics["fields"][display_name] = {
                    "excel_column_index": None,
                    "excel_column_header": None,
                    "raw_cell_value": None,
                    "normalized_value": None,
                }
                continue

            header_value = header_values[index] if index < len(header_values) else None
            raw_value = row_values[index] if index < len(row_values) else None
            if canonical == "maturity_date":
                normalized_value = self._serialize_date(self._parse_date(raw_value))
            else:
                normalized_value = self._coerce_text(raw_value)

            diagnostics["fields"][display_name] = {
                "excel_column_index": index + 1,
                "excel_column_header": self._stringify(header_value),
                "raw_cell_value": self._stringify(raw_value),
                "normalized_value": normalized_value,
            }
        return diagnostics

    def _build_rejection_diagnostic(self, row_values: list[Any], row_number: int, reason: str, *, normalized_row: dict[str, Any] | None, position: dict[str, Any] | None = None) -> dict[str, Any]:
        normalized_cells = {self._normalize_header(key): value for key, value in (normalized_row or {}).items()}
        validation_result = {
            "accepted": position is not None,
            "required_fields": {
                field: {
                    "present": self._coerce_text(self._find_field(normalized_cells, field)) != "",
                    "value": self._stringify(self._find_field(normalized_cells, field)),
                }
                for field in self._REQUIRED_COLUMNS
            },
        }
        return {
            "row": row_number,
            "raw_row_values": [self._stringify(value) for value in row_values],
            "required_fields": list(self._REQUIRED_COLUMNS),
            "validation_result": validation_result,
            "rejection_reason": reason,
        }

    def _normalize_currency(self, value: str) -> str:
        cleaned = self._coerce_text(value).upper()
        if cleaned in {"CRC", "COLONES", "COLON"}:
            return "CRC"
        if cleaned in {"USD", "US$", "$"}:
            return "USD"
        return cleaned

    def _parse_number(self, value: Any) -> Decimal | None:
        if value is None or isinstance(value, bool):
            return None
        if isinstance(value, (int, float, Decimal)):
            return Decimal(str(value))
        text = self._stringify(value).strip()
        if not text:
            return None
        text = text.replace("$", "").replace("€", "").replace("£", "")
        text = text.replace("%", "")
        text = text.strip()
        if not text:
            return None
        if text.startswith("(") and text.endswith(")"):
            text = f"-{text[1:-1]}"
        if "," in text and "." in text:
            if text.rfind(",") > text.rfind("."):
                text = text.replace(".", "").replace(",", ".")
            else:
                text = text.replace(",", "")
        elif "," in text:
            text = text.replace(",", ".")
        try:
            return Decimal(text)
        except Exception:
            return None

    def _parse_date(self, value: Any) -> date | None:
        if value is None:
            return None
        if isinstance(value, datetime):
            parsed_date = value.date()
            return None if self._is_excel_sentinel_date(parsed_date) else parsed_date
        if isinstance(value, date) and not isinstance(value, datetime):
            return None if self._is_excel_sentinel_date(value) else value
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            if math.isfinite(float(value)):
                try:
                    parsed = datetime(1899, 12, 30) + timedelta(days=int(value))
                    parsed_date = parsed.date()
                    return None if self._is_excel_sentinel_date(parsed_date) else parsed_date
                except ValueError:
                    return None
        text = self._stringify(value).strip()
        if not text:
            return None
        if text.lower() in {"1900-01-01", "01/01/1900", "1/1/1900", "1900-1-1"}:
            return None
        try:
            parsed_date = date.fromisoformat(text)
            return None if self._is_excel_sentinel_date(parsed_date) else parsed_date
        except ValueError:
            for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%Y/%m/%d"):
                try:
                    parsed_date = datetime.strptime(text, fmt).date()
                    return None if self._is_excel_sentinel_date(parsed_date) else parsed_date
                except ValueError:
                    continue
            return None

    def _is_excel_sentinel_date(self, value: date) -> bool:
        return value in {date(1899, 12, 31), date(1900, 1, 1)}

    def _serialize_date(self, value: date | None) -> str | None:
        if value is None:
            return None
        if value in {date(1899, 12, 31), date(1900, 1, 1)}:
            return None
        return value.isoformat()

    def _infer_valuation_date(self, path: Path) -> date:
        return date.today()
