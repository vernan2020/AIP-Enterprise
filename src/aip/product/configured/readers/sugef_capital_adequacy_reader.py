from __future__ import annotations

import io
import re
import unicodedata
from calendar import monthrange
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from html.parser import HTMLParser
from typing import Any, Callable
from urllib.parse import urljoin
from urllib.request import Request, urlopen
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from aip.domain.financial_analysis.models import (
    FinancialEntity,
    FinancialStatementLine,
    FinancialStatementType,
    SourceTrace,
)
from aip.product.configured.configuration.configured_source_config import (
    SUGEFFinancialSourceConfig,
)
from aip.product.configured.readers.sugef_public_api_client import SUGEFPublicApiClient


@dataclass(frozen=True, slots=True)
class SUGEFCapitalAdequacyReadResult:
    lines: tuple[FinancialStatementLine, ...]
    source_cutoff: date | None
    source_files: tuple[str, ...]
    diagnostics: tuple[str, ...]


@dataclass(frozen=True, slots=True)
class _DownloadLink:
    href: str
    text: str
    cutoff: date


@dataclass(frozen=True, slots=True)
class _SheetObservation:
    entity: FinancialEntity
    value: Decimal
    row_number: int


class _LinkParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.links: list[tuple[str, str]] = []
        self._href: str | None = None
        self._text: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag.lower() != "a":
            return
        values = dict(attrs)
        href = values.get("href")
        if href:
            self._href = href
            self._text = []

    def handle_data(self, data: str) -> None:
        if self._href is not None:
            self._text.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag.lower() == "a" and self._href is not None:
            self.links.append((self._href, " ".join(self._text).strip()))
            self._href = None
            self._text = []


class SUGEFCapitalAdequacyReader:
    """Lee la suficiencia patrimonial trimestral publicada por SUGEF.

    El valor trimestral se lleva al corte contable solicitado únicamente para
    efectos de calificación, preservando en la trazabilidad la fecha oficial del
    dato. No se interpola ni se recalcula entre publicaciones trimestrales.

    Los archivos descargables de SUGEF pueden venir como una tabla puntual o
    como una hoja histórica. En la variante histórica AIP filtra explícitamente
    el corte oficial seleccionado y admite el acrónimo regulatorio ISPE
    (Indicador de Suficiencia Patrimonial de la Entidad). Si la estructura no
    contiene encabezados reconocibles, se permite una inferencia conservadora:
    debe existir una columna con entidades resolubles en el catálogo oficial y
    una única columna numérica porcentual plausible. Las estructuras ambiguas
    se mantienen como N/D.
    """

    SOURCE_PAGE = "https://www.sugef.fi.cr/reportes/Suficiencia%20Patrimonial.aspx"
    _MONTHS = {
        "ENERO": 1,
        "FEBRERO": 2,
        "MARZO": 3,
        "ABRIL": 4,
        "MAYO": 5,
        "JUNIO": 6,
        "JULIO": 7,
        "AGOSTO": 8,
        "SETIEMBRE": 9,
        "SEPTIEMBRE": 9,
        "OCTUBRE": 10,
        "NOVIEMBRE": 11,
        "DICIEMBRE": 12,
    }

    def __init__(
        self,
        config: SUGEFFinancialSourceConfig,
        *,
        api_client: SUGEFPublicApiClient | None = None,
        fetch_bytes: Callable[[str], bytes] | None = None,
    ) -> None:
        self._config = config
        self._api = api_client or SUGEFPublicApiClient(config)
        self._fetch_bytes = fetch_bytes or self._download

    def read(self, cutoff_date: date) -> SUGEFCapitalAdequacyReadResult:
        try:
            page = self._fetch_bytes(self.SOURCE_PAGE).decode("utf-8", errors="replace")
            links = self._discover_links(page)
            selected = next((link for link in links if link.cutoff <= cutoff_date), None)
            if selected is None:
                return SUGEFCapitalAdequacyReadResult(
                    (),
                    None,
                    (),
                    ("Suficiencia Patrimonial SUGEF: no existe un corte trimestral previo.",),
                )
            workbook_url = urljoin(self.SOURCE_PAGE, selected.href)
            workbook_bytes = self._fetch_bytes(workbook_url)
            entity_index = self._entity_index()
            lines, diagnostics = self._read_workbook(
                workbook_bytes,
                workbook_url=workbook_url,
                source_cutoff=selected.cutoff,
                analysis_cutoff=cutoff_date,
                entity_index=entity_index,
            )
            diagnostics.insert(
                0,
                "Suficiencia Patrimonial SUGEF: último corte trimestral oficial aplicable "
                f"{selected.cutoff:%d/%m/%Y}; usado para el análisis {cutoff_date:%d/%m/%Y}.",
            )
            return SUGEFCapitalAdequacyReadResult(
                lines=lines,
                source_cutoff=selected.cutoff,
                source_files=(workbook_url,),
                diagnostics=tuple(diagnostics),
            )
        except (
            OSError,
            ValueError,
            KeyError,
            InvalidOperation,
            BadZipFile,
            InvalidFileException,
        ) as exc:
            return SUGEFCapitalAdequacyReadResult(
                (),
                None,
                (),
                (f"Suficiencia Patrimonial SUGEF: {type(exc).__name__}: {exc}",),
            )

    def _entity_index(self) -> dict[str, FinancialEntity]:
        response = self._api.list_entities()
        index: dict[str, FinancialEntity] = {}
        for row in response.rows:
            entity_id = self._text(row.get("codigoEntidad"))
            name = self._text(row.get("nombreEntidad")) or self._text(
                row.get("aliasPublicacionEntidad")
            )
            if not entity_id or not name:
                continue
            entity = FinancialEntity(
                entity_id=entity_id,
                name=name,
                category=self._text(row.get("descripcionSector")) or "Sin clasificar",
            )
            aliases = {
                entity_id,
                name,
                self._text(row.get("aliasEntidad")),
                self._text(row.get("aliasPublicacionEntidad")),
            }
            for alias in aliases:
                if alias:
                    index[self._normalize(alias)] = entity
        return index

    @classmethod
    def _discover_links(cls, page: str) -> tuple[_DownloadLink, ...]:
        parser = _LinkParser()
        parser.feed(page)
        links: list[_DownloadLink] = []
        for href, text in parser.links:
            if ".xlsx" not in href.lower():
                continue
            cutoff = cls._cutoff_from_text(f"{text} {href}")
            if cutoff is None:
                continue
            links.append(_DownloadLink(href=href, text=text, cutoff=cutoff))
        return tuple(sorted(links, key=lambda item: item.cutoff, reverse=True))

    @classmethod
    def _cutoff_from_text(cls, value: str) -> date | None:
        normalized = cls._normalize(value)
        for month_name, month in cls._MONTHS.items():
            match = re.search(rf"\b{month_name}\b\D*(20\d{{2}})", normalized)
            if match:
                year = int(match.group(1))
                return date(year, month, monthrange(year, month)[1])
        return None

    @classmethod
    def _read_workbook(
        cls,
        payload: bytes,
        *,
        workbook_url: str,
        source_cutoff: date,
        analysis_cutoff: date,
        entity_index: dict[str, FinancialEntity],
    ) -> tuple[tuple[FinancialStatementLine, ...], list[str]]:
        workbook = load_workbook(io.BytesIO(payload), data_only=True, read_only=True)
        output: list[FinancialStatementLine] = []
        diagnostics: list[str] = []
        unresolved_sheets: list[str] = []
        strategies: set[str] = set()
        try:
            for sheet in workbook.worksheets:
                rows = list(sheet.iter_rows(values_only=True))
                observations, strategy = cls._extract_observations(
                    rows,
                    source_cutoff=source_cutoff,
                    entity_index=entity_index,
                )
                if not observations:
                    unresolved_sheets.append(sheet.title)
                    continue
                strategies.add(strategy)
                for observation in observations:
                    ratio = (
                        observation.value / Decimal("100")
                        if abs(observation.value) > Decimal("1")
                        else observation.value
                    )
                    output.append(
                        FinancialStatementLine(
                            entity=observation.entity,
                            statement_date=analysis_cutoff,
                            statement_type=FinancialStatementType.INDICATORS,
                            account_code="SUGEF:CAPITAL_ADEQUACY",
                            account_name="Suficiencia Patrimonial",
                            amount=ratio,
                            currency="RATIO",
                            trace=SourceTrace(
                                source_name="SUGEF · Suficiencia Patrimonial trimestral",
                                source_url=workbook_url,
                                file_path=f"corte oficial {source_cutoff:%d/%m/%Y}",
                                sheet_name=sheet.title,
                                row_number=observation.row_number,
                            ),
                        )
                    )
        finally:
            workbook.close()

        deduplicated: dict[str, FinancialStatementLine] = {}
        duplicates: set[str] = set()
        for line in output:
            entity_id = line.entity.entity_id
            existing = deduplicated.get(entity_id)
            if existing is None:
                deduplicated[entity_id] = line
                continue
            if existing.amount != line.amount:
                duplicates.add(entity_id)
        if duplicates:
            for entity_id in duplicates:
                deduplicated.pop(entity_id, None)
            diagnostics.append(
                "Suficiencia Patrimonial SUGEF: se excluyeron entidades con más de un valor "
                "distinto para el mismo corte oficial: " + ", ".join(sorted(duplicates)) + "."
            )

        final_output = tuple(deduplicated.values())
        if not final_output:
            if unresolved_sheets:
                diagnostics.append(
                    "Suficiencia Patrimonial SUGEF: no se identificaron observaciones "
                    "inequívocas de entidad, corte e ISPE en las hojas: "
                    f"{', '.join(unresolved_sheets)}."
                )
            diagnostics.append(
                "Suficiencia Patrimonial SUGEF: el XLSX no produjo indicadores utilizables."
            )
        else:
            strategy_text = ", ".join(sorted(strategies))
            diagnostics.append(
                f"Suficiencia Patrimonial SUGEF: {len(final_output)} entidades cargadas desde "
                f"XLSX mediante {strategy_text}."
            )
        return final_output, diagnostics

    @classmethod
    def _extract_observations(
        cls,
        rows: list[tuple[Any, ...]],
        *,
        source_cutoff: date,
        entity_index: dict[str, FinancialEntity],
    ) -> tuple[tuple[_SheetObservation, ...], str]:
        header = cls._header(rows)
        if header is not None:
            header_row, entity_column, value_column, period_column = header
            observations = cls._observations_from_columns(
                rows,
                start_row=header_row + 1,
                entity_column=entity_column,
                value_column=value_column,
                period_column=period_column,
                source_cutoff=source_cutoff,
                entity_index=entity_index,
            )
            if observations:
                return observations, "encabezados explícitos/ISPE"

        inferred = cls._infer_flat_layout(
            rows,
            source_cutoff=source_cutoff,
            entity_index=entity_index,
        )
        if inferred:
            return inferred, "inferencia estructural controlada"
        return (), "sin estructura reconocible"

    @classmethod
    def _observations_from_columns(
        cls,
        rows: list[tuple[Any, ...]],
        *,
        start_row: int,
        entity_column: int,
        value_column: int,
        period_column: int | None,
        source_cutoff: date,
        entity_index: dict[str, FinancialEntity],
    ) -> tuple[_SheetObservation, ...]:
        output: list[_SheetObservation] = []
        for zero_based_row, row in enumerate(rows[start_row:], start=start_row):
            if entity_column >= len(row) or value_column >= len(row):
                continue
            if period_column is not None:
                if period_column >= len(row):
                    continue
                row_cutoff = cls._date_value(row[period_column])
                if row_cutoff != source_cutoff:
                    continue
            entity_name = cls._text(row[entity_column])
            value = cls._decimal(row[value_column])
            if not entity_name or value is None or not cls._plausible_ratio(value):
                continue
            entity = cls._resolve_entity(entity_name, entity_index)
            if entity is None:
                continue
            output.append(
                _SheetObservation(
                    entity=entity,
                    value=value,
                    row_number=zero_based_row + 1,
                )
            )
        return tuple(output)

    @classmethod
    def _header(
        cls,
        rows: list[tuple[Any, ...]],
    ) -> tuple[int, int, int, int | None] | None:
        entity_candidates: list[tuple[int, int]] = []
        value_candidates: list[tuple[int, int]] = []
        period_candidates: list[tuple[int, int]] = []
        for row_index, row in enumerate(rows[:60]):
            for column_index, raw_value in enumerate(row):
                value = cls._normalize(cls._text(raw_value))
                if not value:
                    continue
                if cls._is_entity_header(value):
                    entity_candidates.append((row_index, column_index))
                if cls._is_capital_adequacy_header(value):
                    value_candidates.append((row_index, column_index))
                if cls._is_period_header(value):
                    period_candidates.append((row_index, column_index))

        compatible: list[tuple[int, int, int, int]] = []
        for entity_row, entity_column in entity_candidates:
            for value_row, value_column in value_candidates:
                distance = abs(entity_row - value_row)
                if distance <= 4 and entity_column != value_column:
                    compatible.append(
                        (distance, max(entity_row, value_row), entity_column, value_column)
                    )
        if not compatible:
            return None
        _, header_row, entity_column, value_column = min(compatible)

        period_column: int | None = None
        nearby_periods = [
            (abs(period_row - header_row), period_column_candidate)
            for period_row, period_column_candidate in period_candidates
            if abs(period_row - header_row) <= 4
            and period_column_candidate not in {entity_column, value_column}
        ]
        if nearby_periods:
            period_column = min(nearby_periods)[1]
        return header_row, entity_column, value_column, period_column

    @classmethod
    def _infer_flat_layout(
        cls,
        rows: list[tuple[Any, ...]],
        *,
        source_cutoff: date,
        entity_index: dict[str, FinancialEntity],
    ) -> tuple[_SheetObservation, ...]:
        if not rows:
            return ()
        width = max((len(row) for row in rows), default=0)
        if width < 2:
            return ()

        unique_entity_count = len({entity.entity_id for entity in entity_index.values()})
        minimum_matches = min(3, unique_entity_count)
        if minimum_matches == 0:
            return ()

        entity_matches: dict[int, list[tuple[int, FinancialEntity]]] = {}
        for column in range(width):
            matches: list[tuple[int, FinancialEntity]] = []
            seen: set[str] = set()
            for row_index, row in enumerate(rows):
                if column >= len(row):
                    continue
                entity = cls._resolve_entity(cls._text(row[column]), entity_index)
                if entity is None or entity.entity_id in seen:
                    continue
                seen.add(entity.entity_id)
                matches.append((row_index, entity))
            if len(matches) >= minimum_matches:
                entity_matches[column] = matches
        if not entity_matches:
            return ()

        entity_column = max(entity_matches, key=lambda column: len(entity_matches[column]))
        matched_rows = entity_matches[entity_column]

        period_columns: list[tuple[int, int]] = []
        for column in range(width):
            if column == entity_column:
                continue
            exact_matches = sum(
                1
                for row_index, _ in matched_rows
                if column < len(rows[row_index])
                and cls._date_value(rows[row_index][column]) == source_cutoff
            )
            if exact_matches:
                period_columns.append((exact_matches, column))
        period_column = max(period_columns)[1] if period_columns else None

        candidate_rows = [
            (row_index, entity)
            for row_index, entity in matched_rows
            if period_column is None
            or (
                period_column < len(rows[row_index])
                and cls._date_value(rows[row_index][period_column]) == source_cutoff
            )
        ]
        if len(candidate_rows) < minimum_matches:
            return ()

        value_candidates: list[tuple[int, int, int]] = []
        for column in range(width):
            if column == entity_column or column == period_column:
                continue
            values = [
                cls._decimal(rows[row_index][column])
                for row_index, _ in candidate_rows
                if column < len(rows[row_index])
            ]
            plausible = [
                value for value in values if value is not None and cls._plausible_ratio(value)
            ]
            if len(plausible) < minimum_matches:
                continue
            header_rank = cls._column_header_rank(rows, column)
            value_candidates.append((header_rank, -len(plausible), column))
        if not value_candidates:
            return ()

        value_candidates.sort()
        best_rank = value_candidates[0][:2]
        best_columns = [item[2] for item in value_candidates if item[:2] == best_rank]
        if len(best_columns) != 1:
            return ()
        value_column = best_columns[0]

        output: list[_SheetObservation] = []
        for row_index, entity in candidate_rows:
            if value_column >= len(rows[row_index]):
                continue
            value = cls._decimal(rows[row_index][value_column])
            if value is None or not cls._plausible_ratio(value):
                continue
            output.append(
                _SheetObservation(
                    entity=entity,
                    value=value,
                    row_number=row_index + 1,
                )
            )
        return tuple(output)

    @classmethod
    def _column_header_rank(cls, rows: list[tuple[Any, ...]], column: int) -> int:
        normalized = [
            cls._normalize(cls._text(row[column]))
            for row in rows[:60]
            if column < len(row) and cls._text(row[column])
        ]
        if any(cls._is_capital_adequacy_header(value) for value in normalized):
            return 0
        if any(
            value in {"INDICADOR", "VALOR", "PORCENTAJE", "PORCENTAJE INDICADOR"}
            for value in normalized
        ):
            return 1
        return 2

    @staticmethod
    def _is_entity_header(value: str) -> bool:
        if value in {
            "ENTIDAD",
            "NOMBRE",
            "NOMBRE ENTIDAD",
            "NOMBRE DE LA ENTIDAD",
            "ALIAS ENTIDAD",
            "INSTITUCION",
            "NOMBRE INSTITUCION",
            "NOMBRE DE LA INSTITUCION",
            "ENTIDAD FINANCIERA",
            "INSTITUCION FINANCIERA",
        }:
            return True
        return len(value) <= 80 and ("ENTIDAD" in value or "INSTITUCION" in value)

    @staticmethod
    def _is_capital_adequacy_header(value: str) -> bool:
        if "SUFICIENCIA" in value and "PATRIMONIAL" in value:
            return True
        compact = re.sub(r"[^A-Z0-9]", "", value)
        return compact in {
            "ISP",
            "ISPE",
            "INDICADORDESUFICIENCIAPATRIMONIALDELAENTIDAD",
            "ISFP",
            "SP",
        }

    @staticmethod
    def _is_period_header(value: str) -> bool:
        compact = value.replace(" ", "")
        return value in {
            "FECHA",
            "PERIODO",
            "FECHA CORTE",
            "FECHA DE CORTE",
            "CORTE",
            "MES",
        } or compact in {"FECHACORTE", "FECHADECORTE"}

    @classmethod
    def _resolve_entity(
        cls,
        entity_name: str,
        entity_index: dict[str, FinancialEntity],
    ) -> FinancialEntity | None:
        normalized = cls._normalize(entity_name)
        exact = entity_index.get(normalized)
        if exact is not None:
            return exact

        candidates: dict[str, FinancialEntity] = {}
        for alias, entity in entity_index.items():
            if len(alias) < 5:
                continue
            if alias in normalized or normalized in alias:
                candidates[entity.entity_id] = entity
        if len(candidates) == 1:
            return next(iter(candidates.values()))
        return None

    @classmethod
    def _date_value(cls, value: Any) -> date | None:
        if isinstance(value, datetime):
            parsed = value.date()
            return date(parsed.year, parsed.month, monthrange(parsed.year, parsed.month)[1])
        if isinstance(value, date):
            return date(value.year, value.month, monthrange(value.year, value.month)[1])

        text = cls._text(value)
        if not text:
            return None
        for pattern in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%Y/%m/%d"):
            try:
                parsed = datetime.strptime(text, pattern).date()
            except ValueError:
                continue
            return date(parsed.year, parsed.month, monthrange(parsed.year, parsed.month)[1])
        month_cutoff = cls._cutoff_from_text(text)
        if month_cutoff is not None:
            return month_cutoff
        match = re.fullmatch(r"(20\d{2})[-/](0?[1-9]|1[0-2])", text)
        if match:
            year = int(match.group(1))
            month = int(match.group(2))
            return date(year, month, monthrange(year, month)[1])
        return None

    @staticmethod
    def _plausible_ratio(value: Decimal) -> bool:
        return Decimal("-100") <= value <= Decimal("100")

    @staticmethod
    def _decimal(value: Any) -> Decimal | None:
        if value is None or isinstance(value, bool):
            return None
        if isinstance(value, str):
            text = value.strip().replace("%", "").replace(" ", "")
            if not text:
                return None
            if "," in text and "." in text:
                if text.rfind(",") > text.rfind("."):
                    text = text.replace(".", "").replace(",", ".")
                else:
                    text = text.replace(",", "")
            else:
                text = text.replace(",", ".")
        else:
            text = str(value)
        try:
            return Decimal(text)
        except (InvalidOperation, ValueError):
            return None

    @staticmethod
    def _text(value: Any) -> str:
        return str(value or "").strip()

    @staticmethod
    def _normalize(value: str) -> str:
        decomposed = unicodedata.normalize("NFKD", value)
        return " ".join(
            "".join(character for character in decomposed if not unicodedata.combining(character))
            .upper()
            .split()
        )

    @staticmethod
    def _download(url: str) -> bytes:
        request = Request(
            url,
            headers={
                "User-Agent": "AIP-Enterprise/1.0 SUGEF public-data reader",
                "Accept": "text/html,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*",
            },
        )
        with urlopen(request, timeout=90) as response:
            return response.read()
